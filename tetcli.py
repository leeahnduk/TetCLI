from tetpyclient import RestClient
import tetpyclient
import json
import requests.packages.urllib3
import sys
import os
import argparse
import time
import csv
from columnar import columnar
from time import mktime
from datetime import datetime
import datetime
from argparse import ArgumentParser
from collections import defaultdict
from tqdm import tqdm as progress
import urllib3
import xlsxwriter
from openpyxl import Workbook
import re



CEND = "\33[0m"     #End
CGREEN = "\33[32m"  #Information
CYELLOW = "\33[33m" #Request Input
CRED = "\33[31m"    #Error
URED = "\33[4;31m" 
Cyan = "\33[0;36m"  #Return
BLINK = "\33[5m"
BOLD = "\33[1m"
ITALIC = "\33[3m"
UNDERLINE = "\33[4m"
LBLUE = "\33[1;34m"

# =================================================================================
# feedback: Le Anh Duc - anhdle@cisco.com
# See reason below -- why verify=False param is used
# python3 tetcli.py --url https://10.71.129.30/ --credential Japan_api_credentials.json
# =================================================================================
requests.packages.urllib3.disable_warnings()


parser = argparse.ArgumentParser(description='Tetration Get all sensors')
parser.add_argument('--url', help='Tetration URL', required=True)
parser.add_argument('--credential', help='Path to Tetration json credential file', required=True)
args = parser.parse_args()

# =================================================================================
# Overall
# =================================================================================
def CreateRestClient():
    rc = RestClient(args.url,
                    credentials_file=args.credential, verify=False)
    return rc


# =================================================================================
# Report
# =================================================================================

def ShowAgentProfile(agent):
    """
        Detail of an agent
        """
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Host Name', 'Agent Type', 'Last Check-in', 'Platform', 'Version', 'Scopes']
        data_list = [[agent['host_name'], agent['agent_type'], time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(agent['last_config_fetch_at'])), agent['platform'], agent['current_sw_version'], ','.join(set([y['vrf'] for y in agent['interfaces']]))]]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def selectAgent(sensors):
    # Return UUID for one Sensors that we choose
    print (Cyan + "\nHere are all Software Sensors in your cluster: " + CEND)
    ShowAgents(sensors)
    choice = input('\nSelect which Sensor (Number) above you want to know detail: ')
    return sensors['results'][int(choice)-1]['uuid']


def GetAgentProfile(rc,uuid):
    resp = rc.get('/workload/' + uuid)

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetWorkloadStats(rc,uuid, t0, t1, td):
    #td = 15 * 60 # 15 minutes
    resp = rc.get('/workload/' + uuid + '/stats?t0=' + str(t0) + '&t1=' + str(t1) + '&td=' + str(td))

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowWorkloadStats(stats):
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Time', 'Flow Count', 'received bytes', 'received packets', 'transmitted bytes', 'transmitted packets']
        data_list = [[x['timestamp'], x['result']['flow_count'], x['result']['rx_byte_count'],  x['result']['rx_packet_count'], x['result']['tx_byte_count'], x['result']['tx_packet_count']]for x in stats ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetSwPackages(rc,uuid):
    resp = rc.get('/workload/' + uuid + '/packages')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowSwPackages(packages):
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Name', 'Architecture', 'Publisher', 'Version']
        data_list = [[x['name'], x['architecture'], x['publisher'],  x['version']]for x in packages ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetVul(rc,uuid):
    resp = rc.get('/workload/' + uuid + '/vulnerabilities')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowVul(vuls):
    data_list = []
    headers = ['Package Information', 'CVE ID', 'v2 Score', 'v3 Score', 'v2_severity', 'v2_access_complexity', 'v3_base_severity', 'v3_attack_complexity']
    search_key = 'v3_score'
    for x in vuls: 
        if search_key in x.keys(): data_list.append([x['package_infos'], x['cve_id'], x['v2_score'], x['v3_score'], x['v2_severity'], x['v2_access_complexity'],  x['v3_base_severity'], x['v3_attack_complexity']])
        else: data_list.append([x['package_infos'], x['cve_id'], x['v2_score'], 'None', x['v2_severity'], x['v2_access_complexity'],  'None' , 'None'])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetProc(rc,uuid):
    resp = rc.get('/workload/' + uuid + '/process/list')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowProc(proc):
    data_list = []
    headers = ['PID', 'PPID', 'Proc State', 'Username','CMD', 'Exec_Path', 'Package Name', 'Package Version']
    search_key = 'pkg_info_name'
    for x in proc['ps_row']:
        if search_key in x.keys(): data_list. append([x['pid'], x['ppid'], x['proc_state'],  x['username'], x['cmd'], x['exec_path'], x['pkg_info_name'],  x['pkg_info_version']])
        else: data_list. append([x['pid'], x['ppid'], x['proc_state'], x['username'], x['cmd'], x['exec_path'], 'NA',  'NA'])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetProcTree(rc,uuid):
    payload = {}
    resp = rc.post('/workload/' + uuid + '/process/tree/ids', json_body=json.dumps(payload))

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetProcTreeDetail(rc,uuid, handle):
    payload = {"handle": handle}
    resp = rc.post('/workload/' + uuid + '/process/tree/details', json_body=json.dumps(payload))

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowProcTreeDetail(procDetail):
    data_list = []
    headers = ['PID', 'PPID', 'Proc State', 'Username','CMD', 'Exec_Path', 'Package Name', 'Package Version']
    search_key = 'pkg_info_name'
    for x in procDetail['results']:
        if search_key in x.keys(): data_list. append([x['process_id'], x['parent_process_id'], x['proc_state'],  x['username'], x['command_string'], x['exec_path'], x['pkg_info_name'],  x['pkg_info_version']])
        else: data_list. append([x['process_id'], x['parent_process_id'], x['proc_state'], x['username'], x['command_string'], x['exec_path'], 'NA',  'NA'])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def downloadConvs(rc,appIDs):
    # Download Apps Conversation JSON files from Apps workspace
    apps = []
    limit = int(input ("How many conversation you want to download? "))
    for appID in appIDs:
        print('Downloading app details for '+appID + "into json file")
        versions = GetAppVersions(rc,appID)
        version = int(re.search(r'\d+', GetLatestVersion(versions)).group(0))
        req_payload = {"version": version,
               "limit": limit
               }
        resp = rc.post('/openapi/v1/conversations/%s'%appID, json_body=json.dumps(req_payload))
        if resp.status_code == 200:
            parsed_resp = json.loads(resp.content)
            apps.append(parsed_resp)
    
    with open('all-conversations.json', "w") as config_file:
                json.dump(apps, config_file, indent=4)
                print("all-conversations.json created")



def ShowConversationTet(convs):
    """
        Show All conversation and export to Excel file
        Source IP | Source Filter Name | Destination IP | Destination Filter Name | Protocol | Port | Bytes | Packets
        """
    data_list = []
    headers = ['Source IP', 'Destination IP', 'Protocol', 'Port', 'Bytes', 'Packets']
    listconv = convs[0]
    for x in listconv['results']:
        data_list.append([x['src_ip'], x['dst_ip'], x['protocol'], x['port'], x['byte_count'], x['packet_count']]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)
    with open('conversation.csv', 'w') as csvfile:
        writer = csv.writer(csvfile, delimiter=',')
        writer.writerow(i for i in headers)
        for row in data_list:
            writer.writerow(row)
    
    export_xlsfile = 'Apps_Conversation.xlsx'
    workbook = xlsxwriter.Workbook(export_xlsfile)
    bold = workbook.add_format({'bold': True})
    worksheet = workbook.add_worksheet(name='Apps Conversation')
    header_format = workbook.add_format()
    header_format.set_bg_color('cyan')
    header_format.set_bold()
    header_format.set_font_size(13)
    header_format.set_font_color('black')
    worksheet.set_row(0, None)
    worksheet.write_row(0,0,headers,header_format)
    i=1
    firstline = True
    with open('conversation.csv', 'r') as f:
        for row in csv.reader(f):
            if firstline:    #skip first line
                firstline = False
                continue
            worksheet.write_row(i,0,row)
            i += 1
    worksheet.set_column(0, 0, 20)
    worksheet.set_column(1, 1, 20)
    i =2
    while i < 6:
        worksheet.set_column(i, i, 15)
        i += 1
    workbook.close()
    print ('Writing csv file to %s with %d columns' % (export_xlsfile, len(headers)))
    os.remove('conversation.csv')


def get_inventory(rc, end_point, req_payload):
    '''
    Get the list of inventory items matching the query
    '''

    all_result = []

    resp = rc.post(end_point, json_body=json.dumps(req_payload))
    results = resp.json()

    all_result += results["results"]

    while results.get("offset"):
        # Get the offset ID for page 2
        next_page = results["offset"]
        # Set the offset to page 2
        req_payload["offset"] = next_page

        resp = rc.post(end_point, json_body=json.dumps(req_payload))
        results = resp.json()

        all_result += results["results"]

    return all_result


def get_inventory_cve(rc):

    criticality = int(input("Which CVE Score you want to query your inventory (from 0 to 10): "))
    req_payload = {"filter": {"type": "or",
                              "filters": [{"type": "gt", "field": "host_tags_cvss2", "value": criticality},
                                          {"type": "gt", "field": "host_tags_cvss3", "value": criticality}]}}

#    req_payload = {'filter': {"type": "eq", "field": "ip", "value": "192.168.2.98"}}


    cve_hosts = get_inventory(rc, '/inventory/search', req_payload)

    #print (json.dumps(cve_hosts, indent=4))

    cve_list = []   # store host data with CVE info
    print (BLINK + CRED + 'Processing vulnerabilities data ........ ' + CEND)

    for host in cve_hosts:

        host_uuid = str(host["host_uuid"])

        host_name = host['host_name']

        results = GetVul(rc,host_uuid)

        #print (CYELLOW + 'Gathering CVE data for ' + host_name + " with UUID " + host_uuid + CEND)

        #print (json.dumps(results, indent=4))

        for pkg in results:
            cve_dict = {}
            if "v2_score" in pkg.keys():
                if "v3_score" in pkg.keys():
                    if (int(pkg["v2_score"]) > criticality) or int(pkg["v3_score"]) > criticality :
                        cve_dict["IP"] = host["ip"]
                        cve_dict["Hostname"] = host["host_name"]
                        cve_dict["OS"] = host["os"]
                        cve_dict["Version"] = host["os_version"]
                        cve_dict["Package Info"] = pkg["package_infos"]
                        cve_dict["Scope"] = host["tags_scope_name"]
                        cve_dict["CVE ID"] = pkg["cve_id"]
                   
                        cve_dict["CVE v2 Score"] = pkg["v2_score"]
                        cve_dict["CVE v2 Severity"] = pkg["v2_severity"]
                        cve_dict["CVE v2 access vector"] = pkg["v2_access_vector"]
                        cve_dict["CVE v2 access complexity"] = pkg["v2_access_complexity"]

                        cve_dict["CVE v3 Score"] = pkg["v3_score"]
                        cve_dict["CVE v3 Severity"] = pkg["v3_base_severity"]
                        cve_dict["CVE v3 attack vector"] = pkg["v3_attack_vector"]
                        cve_dict["CVE v3 attack complexity"] = pkg["v3_attack_complexity"]
                        cve_dict["CVE v3 availability impact"] = pkg["v3_availability_impact"]
                else:
                    if (int(pkg["v2_score"]) > criticality):
                        cve_dict["IP"] = host["ip"]
                        cve_dict["Hostname"] = host["host_name"]
                        cve_dict["OS"] = host["os"]
                        cve_dict["Version"] = host["os_version"]
                        cve_dict["Package Info"] = pkg["package_infos"]
                        cve_dict["Scope"] = host["tags_scope_name"]
                        cve_dict["CVE ID"] = pkg["cve_id"]
                   
                        cve_dict["CVE v2 Score"] = pkg["v2_score"]
                        cve_dict["CVE v2 Severity"] = pkg["v2_severity"]
                        cve_dict["CVE v2 access vector"] = pkg["v2_access_vector"]
                        cve_dict["CVE v2 access complexity"] = pkg["v2_access_complexity"]
            else:
                if (int(pkg["v3_score"]) > criticality):
                    cve_dict["IP"] = host["ip"]
                    cve_dict["Hostname"] = host["host_name"]
                    cve_dict["OS"] = host["os"]
                    cve_dict["Version"] = host["os_version"]
                    cve_dict["Package Info"] = pkg["package_infos"]
                    cve_dict["Scope"] = host["tags_scope_name"]
                    cve_dict["CVE ID"] = pkg["cve_id"]
               
                    cve_dict["CVE v3 Score"] = pkg["v3_score"]
                    cve_dict["CVE v3 Severity"] = pkg["v3_base_severity"]
                    cve_dict["CVE v3 attack vector"] = pkg["v3_attack_vector"]
                    cve_dict["CVE v3 attack complexity"] = pkg["v3_attack_complexity"]
                    cve_dict["CVE v3 availability impact"] = pkg["v3_availability_impact"]

            cve_list.append(cve_dict)

            cve_list_final = []
            for string in cve_list:
                if (string != ""): cve_list_final.append(string)

    # specify csv file for exporting
    export_xlsfile = 'cve_hosts_final.xlsx'
    export_csvfile = 'cve_hosts_final.csv'
    temp_csv = 'cve_hosts.csv'

    # specify csv header fields
    csv_header = ["IP", "Hostname", "OS", "Version", "Package Info", "Scope", "CVE ID", 
                  "CVE v2 Score", "CVE v2 Severity", "CVE v2 access vector", "CVE v2 access complexity",
                  "CVE v3 Score", "CVE v3 Severity", "CVE v3 attack vector", "CVE v3 attack complexity", "CVE v3 availability impact"]

    
    # Export file in csv format
    with open(temp_csv, 'w+') as f:
        writer = csv.DictWriter(f, csv_header, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        for row in cve_list_final:
            writer.writerow(row)

    with open('cve_hosts.csv') as infile, open('cve_hosts_final.csv', 'w', newline='') as output:
         writer = csv.writer(output)
         for row in csv.reader(infile):
             if any(field.strip() for field in row):
                 writer.writerow(row)

    workbook = xlsxwriter.Workbook(export_xlsfile)
    bold = workbook.add_format({'bold': True})
    worksheet = workbook.add_worksheet(name='CVE Report')
    header_format = workbook.add_format()
    header_format.set_bg_color('cyan')
    header_format.set_bold()
    header_format.set_text_wrap()
    header_format.set_font_size(13)
    header_format.set_font_color('black')
    cell_format = workbook.add_format()
    cell_format.set_text_wrap()
    worksheet.set_row(0, None)
    worksheet.write_row(0,0,csv_header,header_format)
    i=1
    firstline = True
    with open('cve_hosts_final.csv', 'r') as f:
        for row in csv.reader(f):
            if firstline:    #skip first line
                firstline = False
                continue
            worksheet.write_row(i,0,row)
            i += 1
    worksheet.set_column(0, 0, 15)
    worksheet.set_column(1, 1, 15)
    worksheet.set_column(2, 2, 15,cell_format)
    worksheet.set_column(4, 4, 30,cell_format)
    worksheet.set_column(5, 5, 30,cell_format)
    i =6
    while i < 16:
        worksheet.set_column(i, i, 15)
        i += 1
    workbook.close()
    print ('Writing csv file to %s with %d columns' % (export_xlsfile, len(csv_header)))
    os.remove(temp_csv)
    os.remove(export_csvfile)


def get_inventory_flow(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to get inventory statistic: ')
    scope_name = scopes[int(choice)-1]['name']
    subnet = input (CYELLOW + "Which subnet (X.X.X.X/Y) of inventory you want to query: " +CEND)
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),int(from_day),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),int(to_day),0,0).timestamp())
    # Query inventories in the scope
    req_payload = {
    "filter":
            {
                "type": "subnet",
                "field": "ip",
                "value": subnet
            },
    "scopeName": scope_name }

#    req_payload = {'filter': {"type": "eq", "field": "ip", "value": "192.168.2.98"}}

    hosts_in_scope = get_inventory(rc, '/inventory/search', req_payload)

    talkative_list = []   # store host data with bytes
    print (BLINK + CRED +'Processing flow data ........ '+ CEND)

    for host in hosts_in_scope:
        req_endpoint = '/inventory/' + str(host["ip"]) + '-' + str(host["vrf_id"] + '/stats?t0=' + str(t0) +'&t1='+str(t1)+'&td=day')
        #print (req_endpoint)
        
        results = rc.get(req_endpoint).json()

        #print ('Getting conversation data from ' + req_endpoint)
        for x in results:
            stats_dict = {}
            stats_dict["Hostname"] = host["host_name"]
            stats_dict["IP"] = host["ip"]
            stats_dict["Timestamp"] = x["timestamp"]
            stats_dict["OS"] = host["os"]
            stats_dict["OS Version"] = host["os_version"]
            stats_dict["MAC Address"] = host["iface_mac"]
            stats_dict["Received Bytes"] = x["result"]["rx_byte_count"]
            stats_dict["Transmited Bytes"] = x["result"]["tx_byte_count"]
            stats_dict["Total Flows"] = x["result"]["flow_count"]
            stats_dict["Received Packets"] = x["result"]["rx_packet_count"]
            stats_dict["Transmited Packets"] = x["result"]["tx_packet_count"]
        talkative_list.append(stats_dict)

    # specify csv file for exporting
    export_csvfile = 'stats_hosts.xlsx'

    # specify csv header fields
    csv_header = ["Hostname", "IP", "Timestamp", "OS", "OS Version", "MAC Address", "Received Bytes", "Transmited Bytes",
                  "Total Flows", "Received Packets", "Transmited Packets"]

    workbook = xlsxwriter.Workbook(export_csvfile)
    bold = workbook.add_format({'bold': True})
    worksheet = workbook.add_worksheet(name='Subnet Top talkers')
    cell_format = workbook.add_format()
    cell_format.set_bg_color('cyan')
    cell_format.set_bold()
    cell_format.set_font_color('black')
    worksheet.set_row(0, None)
    worksheet.write_row(0,0,csv_header,cell_format)
    i=1
    for row in talkative_list:
        #print (row.values())
        worksheet.write_row(i,0,row.values())
        i+=1
    worksheet.set_column(0, 0, 18)
    worksheet.set_column(1, 1, 15)
    worksheet.set_column(2, 2, 22)
    i =3
    while i < 12:
        worksheet.set_column(i, i, 15)
        i += 1
    workbook.close()

    print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))

def get_flow_topTalkers(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to query Top Talkers: ')
    scope_name = scopes[int(choice)-1]['name']
    threshold = input('\nHow many top talkers you want to query (Max is 1000): ')
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
    #dimensions = GetFlowDimensions(rc)
    metrics = GetFlowMetrics(rc)
    print (Cyan + "Here are the available metrics: \n" + json.dumps(metrics, indent=4, sort_keys=True) + CEND)
    metric = input (Cyan + "which one you want to query? (copy and paste here): " +  CEND)
    #print (Cyan + "Here are the available dimensions: \n" + json.dumps(metrics, indent=4, sort_keys=True))
    #dimension = input (Cyan + "which one you want to query? (copy and paste here) ")
    req_payload = {
    "t0": t0,    
    "t1": t1,    
    "dimension": "src_address",
    "metric": metric,
    #"filter": {"type": "eq", "field": "src_address", "value": "172.29.203.193"},  #optional
    "threshold": int(threshold),
    "scopeName": scope_name
    }


    resp = rc.post('/flowsearch/topn',
               json_body=json.dumps(req_payload))

    #print (json.dumps(cve_hosts, indent=4))
    if resp.status_code != 200:
        print(URED + "Failed to retrieve TopN")
        print(resp.status_code)
        print(resp.text)
    else:
        topN = resp.json()
        print (json.dumps(topN, indent=4))
        topN_list = []   # store TopN data
        print (json.dumps(topN[0]['result'], indent=4))
        for top in topN[0]['result']:
            topN_dict = {}
            topN_dict["Source Address"] = top["src_address"]
            topN_dict[metric] = top[metric]
            topN_list.append(topN_dict)

        # specify csv file for exporting
        export_csvfile = 'topTalkerReport.xlsx'

        # specify csv header fields
        csv_header = ["Source Address", metric]

        
        
        workbook = xlsxwriter.Workbook(export_csvfile)
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet(name='Top Source Address')
        cell_format = workbook.add_format()
        cell_format.set_bg_color('cyan')
        cell_format.set_bold()
        cell_format.set_font_color('black')
        worksheet.set_row(0, None)
        worksheet.write_row(0,0,csv_header,cell_format)
        i=1
        for row in topN_list:
            #print (row.values())
            worksheet.write_row(i,0,row.values())
            i+=1
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        workbook.close()

        print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))

def get_flow_topDest(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to query Top Destination: ')
    scope_name = scopes[int(choice)-1]['name']
    threshold = input('\nHow many top Destination you want to query (Max is 1000): ')
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
    #dimensions = GetFlowDimensions(rc)
    metrics = GetFlowMetrics(rc)
    print (Cyan + "Here are the available metrics: \n" + json.dumps(metrics, indent=4, sort_keys=True) + CEND)
    metric = input (Cyan + "which one you want to query? (copy and paste here): " +  CEND)
    #print (Cyan + "Here are the available dimensions: \n" + json.dumps(metrics, indent=4, sort_keys=True))
    #dimension = input (Cyan + "which one you want to query? (copy and paste here) ")
    req_payload = {
    "t0": t0,    
    "t1": t1,    
    "dimension": "dst_address",
    "metric": metric,
    #"filter": {"type": "eq", "field": "src_address", "value": "172.29.203.193"},  #optional
    "threshold": int(threshold),
    "scopeName": scope_name
    }


    resp = rc.post('/flowsearch/topn',
               json_body=json.dumps(req_payload))

    #print (json.dumps(cve_hosts, indent=4))
    if resp.status_code != 200:
        print(URED + "Failed to retrieve TopN")
        print(resp.status_code)
        print(resp.text)
    else:
        topN = resp.json()
        print (json.dumps(topN, indent=4))
        topN_list = []   # store TopN data
        print (json.dumps(topN[0]['result'], indent=4))
        for top in topN[0]['result']:
            topN_dict = {}
            topN_dict["Destination Address"] = top["dst_address"]
            topN_dict[metric] = top[metric]
            topN_list.append(topN_dict)

        # specify csv file for exporting
        export_csvfile = 'topDestinationReport.xlsx'

        # specify csv header fields
        csv_header = ["Destination Address", metric]

        
        
        workbook = xlsxwriter.Workbook(export_csvfile)
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet(name='Top Destination Address')
        cell_format = workbook.add_format()
        cell_format.set_bg_color('cyan')
        cell_format.set_bold()
        cell_format.set_font_color('black')
        worksheet.set_row(0, None)
        worksheet.write_row(0,0,csv_header,cell_format)
        i=1
        for row in topN_list:
            #print (row.values())
            worksheet.write_row(i,0,row.values())
            i+=1
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        workbook.close()

        print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))

def get_flow_topDestService(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to query Top Server Service: ')
    scope_name = scopes[int(choice)-1]['name']
    threshold = input('\nHow many Top Server Service you want to query (Max is 1000): ')
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
    #dimensions = GetFlowDimensions(rc)
    metrics = GetFlowMetrics(rc)
    print (Cyan + "Here are the available metrics: \n" + json.dumps(metrics, indent=4, sort_keys=True) + CEND)
    metric = input (Cyan + "which one you want to query? (copy and paste here): " +  CEND)
    #print (Cyan + "Here are the available dimensions: \n" + json.dumps(metrics, indent=4, sort_keys=True))
    #dimension = input (Cyan + "which one you want to query? (copy and paste here) ")
    req_payload = {
    "t0": t0,    
    "t1": t1,    
    "dimension": "dst_port",
    "metric": metric,
    #"filter": {"type": "eq", "field": "src_address", "value": "172.29.203.193"},  #optional
    "threshold": int(threshold),
    "scopeName": scope_name
    }


    resp = rc.post('/flowsearch/topn',
               json_body=json.dumps(req_payload))

    #print (json.dumps(cve_hosts, indent=4))
    if resp.status_code != 200:
        print(URED + "Failed to retrieve TopN")
        print(resp.status_code)
        print(resp.text)
    else:
        topN = resp.json()
        print (json.dumps(topN, indent=4))
        topN_list = []   # store TopN data
        print (json.dumps(topN[0]['result'], indent=4))
        for top in topN[0]['result']:
            topN_dict = {}
            topN_dict["Destination Service"] = top["dst_port"]
            topN_dict[metric] = top[metric]
            topN_list.append(topN_dict)

        # specify csv file for exporting
        export_csvfile = 'topDestinationPort.xlsx'

        # specify csv header fields
        csv_header = ["Destination Service", metric]

        
        
        workbook = xlsxwriter.Workbook(export_csvfile)
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet(name='Top Destination Service')
        cell_format = workbook.add_format()
        cell_format.set_bg_color('cyan')
        cell_format.set_bold()
        cell_format.set_font_color('black')
        worksheet.set_row(0, None)
        worksheet.write_row(0,0,csv_header,cell_format)
        i=1
        for row in topN_list:
            #print (row.values())
            worksheet.write_row(i,0,row.values())
            i+=1
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        workbook.close()
        print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))

def get_flow_topSrcService(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to query Top Client Service: ')
    scope_name = scopes[int(choice)-1]['name']
    threshold = input('\nHow many Top Client Service you want to query (Max is 1000): ')
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
    #dimensions = GetFlowDimensions(rc)
    metrics = GetFlowMetrics(rc)
    print (Cyan + "Here are the available metrics: \n" + json.dumps(metrics, indent=4, sort_keys=True) + CEND)
    metric = input (Cyan + "which one you want to query? (copy and paste here): " +  CEND)
    #print (Cyan + "Here are the available dimensions: \n" + json.dumps(metrics, indent=4, sort_keys=True))
    #dimension = input (Cyan + "which one you want to query? (copy and paste here) ")
    req_payload = {
    "t0": t0,    
    "t1": t1,    
    "dimension": "src_port",
    "metric": metric,
    #"filter": {"type": "eq", "field": "src_address", "value": "172.29.203.193"},  #optional
    "threshold": int(threshold),
    "scopeName": scope_name
    }


    resp = rc.post('/flowsearch/topn',
               json_body=json.dumps(req_payload))

    #print (json.dumps(cve_hosts, indent=4))
    if resp.status_code != 200:
        print(URED + "Failed to retrieve TopN")
        print(resp.status_code)
        print(resp.text)
    else:
        topN = resp.json()
        #print (json.dumps(topN, indent=4))
        topN_list = []   # store TopN data
        print (json.dumps(topN[0]['result'], indent=4))
        for top in topN[0]['result']:
            topN_dict = {}
            topN_dict["Source Service"] = top["src_port"]
            topN_dict[metric] = top[metric]
            topN_list.append(topN_dict)

        # specify csv file for exporting
        export_csvfile = 'topSrcPort.xlsx'

        # specify csv header fields
        csv_header = ["Source Service", metric]

        
        workbook = xlsxwriter.Workbook(export_csvfile)
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet(name='Top Source Service')
        cell_format = workbook.add_format()
        cell_format.set_bg_color('cyan')
        cell_format.set_bold()
        cell_format.set_font_color('black')
        worksheet.set_row(0, None)
        worksheet.write_row(0,0,csv_header,cell_format)
        i=1
        for row in topN_list:
            #print (row.values())
            worksheet.write_row(i,0,row.values())
            i+=1
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        workbook.close()
        print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))


# =================================================================================
# Convert
# =================================================================================

def filterToString(invfilter):
    if 'filters' in invfilter.keys():
        query=[]
        for x in invfilter['filters']:
            if 'filters' in x.keys():
                query.append(filterToString(x))
            elif 'filter' in x.keys():
                query.append(x['type'] + filterToString(x['filter']))
            else:
                query.append(x['field'].replace('user_','*')+ ' '+ x['type'] + ' '+ str(x['value']))
        operator = ' '+invfilter['type']+' '
        return '('+operator.join(query)+')'
    else:
        return invfilter['field']+ ' '+ invfilter['type'] + ' '+ str(invfilter['value'])

def selectTetApps(apps):
    # Return App IDa for one or many Tetration Apps that we choose
    print (Cyan + "\nHere are all Application workspaces in your cluster: " + CEND)
    ShowApps(apps)
    choice = input('\nSelect which Tetration Apps (Number, Number) above you want to download polices: ')

    choice = choice.split(',')
    appIDs = []
    for app in choice:
        if '-' in app:
            for app in range(int(app.split('-')[0])-1,int(app.split('-')[1])):
                appIDs.append(resp.json()[int(app)-1]['id'])
        else:
            appIDs.append(apps[int(app)-1]['id'])
    return appIDs

def downloadPolicies(rc,appIDs):
    # Download Policies JSON files from Apps workspace
    apps = []
    for appID in appIDs:
        print('Downloading app details for '+appID + "into json file")
        apps.append(rc.get('/openapi/v1/applications/%s/details'%appID).json())
        #json_object = json.load(apps)
    for app in apps:
        with open('./'+app['name'].replace('/','-')+'.json', "w") as config_file:
            json.dump(apps, config_file, indent=4)
            print(app['name'].replace('/','-')+".json created")
    return apps

def convApps2xls(rc):
    AllApps = GetApps(rc)
    scopes = GetApplicationScopes(rc)
    apps = []
    appIDs = selectTetApps(AllApps)
    apps.append(downloadPolicies(rc, appIDs))
    print (BLINK + CRED +'Processing Application data ........ '+ CEND)
    #print (json.dumps(apps, indent=4))

    # Load in the IANA Protocols
    protocols = {}
    try: 
        with open('protocol-numbers-1.csv') as protocol_file:
            reader = csv.DictReader(protocol_file)
            for row in reader:
                protocols[row['Decimal']]=row
    except IOError:
        print('%% Could not load protocols file')
        return
    except ValueError:
        print('Could not load improperly formatted protocols file')
        return
    
    for app in apps[0]:
        workbook = xlsxwriter.Workbook(app['name'].replace('/','-')+'.xlsx')
        bold = workbook.add_format({'bold': True})

        if 'clusters' in app.keys():
            worksheet = workbook.add_worksheet(name='App Servers')
            cell_format = workbook.add_format()
            cell_format.set_bg_color('cyan')
            cell_format.set_bold()
            cell_format.set_font_color('black')
            worksheet.set_row(0, None)
            worksheet.write_row(0,0,['Hostname','IP','Cluster Membership'],cell_format)
            i=1
            clusters = app['clusters']
            for cluster in clusters:
                hosts = []
                for node in cluster['nodes']:
                    hosts.append(node['name'])
                    worksheet.write_row(i,0,[node['name'],node['ip'],cluster['name']])
                    i+=1
            worksheet.set_column(0, 0, 30)
            worksheet.set_column(1, 1, 15)
            worksheet.set_column(2, 2, 30)

        if 'inventory_filters' in app.keys():
            i=1
            worksheet = workbook.add_worksheet(name='External Groups')
            cell_format = workbook.add_format()
            cell_format.set_text_wrap()
            header_format = workbook.add_format()
            header_format.set_bg_color('cyan')
            header_format.set_bold()
            header_format.set_font_color('black')
            worksheet.set_row(0, None)
            worksheet.write_row(0,0,['Inventory Filter Name', 'IP Addresses', 'Filter Definition'],header_format)
            worksheet.set_column(0, 0, 30)
            worksheet.set_column(1, 1, 60, cell_format)
            worksheet.set_column(2, 2, 50, cell_format)

            filters = app['inventory_filters']
            for invfilter in filters:
                #print (json.dumps(invfilter, indent=4))
                ipSet = resolveFilter(rc, invfilter)
                #print (ipSet)
                worksheet.write_row(i,0,[invfilter['name'], str(ipSet), filterToString(invfilter['query'])])
                i+=1

        if 'default_policies' in app.keys():
            i=1
            worksheet = workbook.add_worksheet(name='Policies')
            header_format = workbook.add_format()
            header_format.set_bg_color('cyan')
            header_format.set_bold()
            header_format.set_font_color('black')
            worksheet.set_row(0, None)
            worksheet.write_row(0,0,['Consumer Group','Provider Group','Services'],header_format)
            worksheet.set_column(0, 0, 30)
            worksheet.set_column(1, 1, 30)
            worksheet.set_column(2, 2, 30)

            policies = app['default_policies']
            for policy in policies:
                pols = {}
                for rule in policy['l4_params']:
                    if 'port' in rule:
                        if rule['port'][0] == rule['port'][1]:
                            port = str(rule['port'][0])
                        else:
                            port = str(rule['port'][0]) + '-' + str(rule['port'][1])
                    else:
                        port = None

                    if port == None:
                        try:
                            pols[protocols[str(rule['proto'])]['Keyword']] = []
                        except:
                            pols['PROTO-'+str(rule['proto'])]=[]
                    elif protocols[str(rule['proto'])]['Keyword'] in pols.keys():
                        pols[protocols[str(rule['proto'])]['Keyword']].append(port)
                    else:
                        pols[protocols[str(rule['proto'])]['Keyword']] = [port]

                policy_list = []
                for key, val in pols.items():
                    #print(key,val)
                    if len(val)>0:
                        policy_list.append('{}={}'.format(key,', '.join(val)))
                    else:
                        policy_list.append(key)
                        
                worksheet.write_row(i,0,[policy["consumer_filter_name"],policy["provider_filter_name"],'; '.join(policy_list)])
                i+=1
        
        workbook.close()
        print (app['name'].replace('/','-')+'.xlsx created for policies conversion to CSV')


def resolveFilter(rc, filters):# return all IP and hosts for a specific filters
    ipSet = []
    body = json.dumps({'filter':filters['query']})
   
    resp = rc.post('/inventory/search',json_body=body)
    if resp:
        ips = resp.json()
        for i in ips['results']:
            ipSet.append(i['ip'])

    return ipSet

def GetInvFromApps(apps):
    for app in apps[0]:
        if 'inventory_filters' in app.keys():
            return app['inventory_filters']
        else: print("CRED + There's no inventory filters in the apps")

def convApps2asa(rc):
    AllApps = GetApps(rc)
    scopes = GetApplicationScopes(rc)
    apps = []
    appIDs = selectTetApps(AllApps)
    apps.append(downloadPolicies(rc, appIDs))
    def_policies = getDefaultDetail(rc,str(appIDs[0]))

    # Load in the IANA Protocols
    protocols = {}
    try: 
        with open('protocol-numbers-1.csv') as protocol_file:
            reader = csv.DictReader(protocol_file)
            for row in reader:
                protocols[row['Decimal']]=row
    except IOError:
        print('%% Could not load protocols file')
        return
    except ValueError:
        print('Could not load improperly formatted protocols file')
        return
    
    # Load in ASA known ports
    ports = {}
    try:
        with open('asa_ports.csv') as protocol_file:
            reader = csv.DictReader(protocol_file)
            for row in reader:
                ports[row['Port']]=row
    except IOError:
        print ('%% Could not load protocols file')
        return
    except ValueError:
        print ('Could not load improperly formatted protocols file')
        return

    print('\nASA ACL Config\n---------------------------------------\n\n')
    #Process nodes and output information to ASA Objects
    file1 = open("ACL_config.txt","w")
    for app in apps[0]:
        if 'clusters' in app.keys():
            clusters = GetClusters(rc,str(appIDs[0]))
            for cluster in clusters:
                print ("object network " + cluster['name'].replace(' ','_'))
                file1.write("object network " + cluster['name'].replace(' ','_') + "\n")
                ClustersipSet = resolveFilter(rc, cluster)
                for ip in ClustersipSet:
                    print ("  host " + ip)
                    file1.write("  host " + ip + "\n")
        if 'inventory_filters' in app.keys():
            filters = GetInvFromApps(apps)
            for invfilter in filters:
                if invfilter['name'] != 'Default':
                    print ("object network " + invfilter['name'].replace(' ','_'))
                    file1.write("object network " + invfilter['name'].replace(' ','_')+ "\n")
                    FiltersipSet = resolveFilter(rc, invfilter)
                    for ip in FiltersipSet:
                        print ("  host " + ip)
                        file1.write("  host " + ip + "\n")

    print ('!')
    file1.write('! \n')

    #Process policies and output information as ASA ACL Lines
    for policy in def_policies:
        for param in policy['l4_params']:
            l4params = []
            if param['proto'] == 1: l4params.append({'port_min': 'NA' ,'port_max': 'NA','proto':param['proto']})
            else: l4params.append({'port_min':param['port'][0],'port_max':param['port'][1],'proto':param['proto']})
 
        for rule in l4params:
            if policy['consumer_filter_id'] != policy['provider_filter_id']:
                if rule['proto'] == 1:
                    print ("access-list ACL_IN extended permit " + protocols[str(rule['proto'])]['Keyword'] + ((" object " + policy['consumer_filter']['name'].replace(' ','_')) if policy['provider_filter']['name'] != 'Default' else " any") + ((" object " + policy['provider_filter']['name'].replace(' ','_')) if policy['provider_filter']['name'] != 'Default' else " any"))
                    file1.write("access-list ACL_IN extended permit " + protocols[str(rule['proto'])]['Keyword'] + ((" object " + policy['consumer_filter']['name'].replace(' ','_')) if policy['provider_filter']['name'] != 'Default' else " any") + ((" object " + policy['provider_filter']['name'].replace(' ','_')) if policy['provider_filter']['name'] != 'Default' else " any") + '\n')               
                elif (rule['proto'] == 6) or (rule['proto'] == 17):
                    if rule['port_min'] == rule['port_max']:
                        if (str(rule['port_min']) in ports.keys()) and (ports[str(rule['port_min'])]['Proto'] == protocols[str(rule['proto'])]['Keyword'] or ports[str(rule['port_min'])]['Proto'] == 'TCP, UDP'):
                            port = ports[str(rule['port_min'])]['Name']
                        else:
                            port = rule['port_min']
                        print ("access-list ACL_IN extended permit " + protocols[str(rule['proto'])]['Keyword'] + ((" object " + policy['consumer_filter']['name'].replace(' ','_')) if policy['consumer_filter']['name'] != 'Default' else " any") + ((" object " + policy['provider_filter']['name'].replace(' ','_')) if policy['provider_filter']['name'] != 'Default' else " any") + " eq " + str(port))
                        file1.write("access-list ACL_IN extended permit " + protocols[str(rule['proto'])]['Keyword'] + ((" object " + policy['consumer_filter']['name'].replace(' ','_')) if policy['consumer_filter']['name'] != 'Default' else " any") + ((" object " + policy['provider_filter']['name'].replace(' ','_')) if policy['provider_filter']['name'] != 'Default' else " any") + " eq " + str(port) + "\n")
                    else:
                        print ("access-list ACL_IN extended permit " + protocols[str(rule['proto'])]['Keyword'] + ((" object " + policy['consumer_filter']['name'].replace(' ','_')) if policy['consumer_filter']['name'] != 'Default' else " any") + ((" object " + policy['provider_filter']['name'].replace(' ','_')) if policy['provider_filter']['name'] != 'Default' else " any") + " range " + str(rule['port_min']) + "-" + str(rule['port_max']))
                        file1.write("access-list ACL_IN extended permit " + protocols[str(rule['proto'])]['Keyword'] + ((" object " + policy['consumer_filter']['name'].replace(' ','_')) if policy['consumer_filter']['name'] != 'Default' else " any") + ((" object " + policy['provider_filter']['name'].replace(' ','_')) if policy['provider_filter']['name'] != 'Default' else " any") + " range " + str(rule['port_min']) + "-" + str(rule['port_max']) + "\n")
    print ("access-list ACL_IN extended deny ip any any\n!\n\n")
    file1.write("access-list ACL_IN extended deny ip any any\n!\n\n")
    file1.close()
    print (CYELLOW + "ACL Config File: ACL_config.txt created" + CEND)


def convApps2n9k(rc):
    AllApps = GetApps(rc)
    scopes = GetApplicationScopes(rc)
    apps = []
    appIDs = selectTetApps(AllApps)
    apps.append(downloadPolicies(rc, appIDs))
    def_policies = getDefaultDetail(rc,str(appIDs[0]))
    #print ("Default Policies: \n" +json.dumps(def_policies, indent=4))
    abs_policies = getAbsoluteDetail(rc,str(appIDs[0]))
    #print ("Absolute Policies: \n" + json.dumps(abs_policies, indent=4))


    # Load in the IANA Protocols
    protocols = {}
    try: 
        with open('protocol-numbers-1.csv') as protocol_file:
            reader = csv.DictReader(protocol_file)
            for row in reader:
                protocols[row['Decimal']]=row
    except IOError:
        print('%% Could not load protocols file')
        return
    except ValueError:
        print('Could not load improperly formatted protocols file')
        return
    
    # Load in N9k known ports
    ports = {}
    print('\nN9k ACL Config\n---------------------------------------\n\n')
    #Process nodes and output information to N9k Objects
    file1 = open("ACL_config.txt","w")
    print ('ip access-list tet-acl')
    file1.write('ip access-list tet-acl \n')
   
    #Process policies and output information as N9k ACL Lines
    for policy in def_policies:
        #print ("Policy: \n" + json.dumps(policy, indent=4))
        for param in policy['l4_params']:
            #print ("L4 Param: \n" + json.dumps(param, indent=4))
            l4params = []
            if param['proto'] == 1: l4params.append({'port_min': 'NA' ,'port_max': 'NA','proto':param['proto']})
            else: l4params.append({'port_min':param['port'][0],'port_max':param['port'][1],'proto':param['proto']})
            #if policy['consumer_filter']['name'] == 'Default' and policy['provider_filter']['name'] != 'Default':
        #print ("L4 Params: \n" + json.dumps(l4params, indent=4)) 
        for rule in l4params:
            if policy['consumer_filter_id'] != policy['provider_filter_id']:
                if rule['proto'] == 1:
                    for app in apps[0]:
                        if 'clusters' in app.keys():
                            clusters = GetClusters(rc,str(appIDs[0]))
                            for cluster in clusters:
                                if policy['provider_filter']['name'] == cluster['name']:
                                    ProvipSet = resolveFilter(rc, cluster)
                                if policy['consumer_filter']['name'] == cluster['name']:
                                    ConsipSet = resolveFilter(rc, cluster)
                        if 'inventory_filters' in app.keys():
                            filters = GetInvFromApps(apps)
                            for invfilter in filters:
                                if invfilter['name'] != 'Default':
                                    if policy['provider_filter']['name'] == invfilter['name']:
                                        ProvipSet = resolveFilter(rc, invfilter)
                                    if policy['consumer_filter']['name'] == invfilter['name']:
                                        ConsipSet = resolveFilter(rc, invfilter)
                        for a in ConsipSet:
                            for b in ProvipSet: 
                                if a != b:
                                    print ("\t permit " + protocols[str(rule['proto'])]['Keyword'] + " host " + (a if policy['provider_filter']['name'] != 'Default' else " any") + " host " + (b if policy['provider_filter']['name'] != 'Default' else " any"))
                                    file1.write("\t permit " + protocols[str(rule['proto'])]['Keyword'] + " host " + (a if policy['provider_filter']['name'] != 'Default' else " any") + " host " + (b if policy['provider_filter']['name'] != 'Default' else " any\n"))           
                elif (rule['proto'] == 6) or (rule['proto'] == 17):
                    for app in apps[0]:
                        if 'clusters' in app.keys():
                            clusters = GetClusters(rc,str(appIDs[0]))
                            for cluster in clusters:
                                if policy['provider_filter']['name'] == cluster['name']:
                                    ProvipSet = resolveFilter(rc, cluster)
                                if policy['consumer_filter']['name'] == cluster['name']:
                                    ConsipSet = resolveFilter(rc, cluster)
                        if 'inventory_filters' in app.keys():
                            filters = GetInvFromApps(apps)
                            for invfilter in filters:
                                if invfilter['name'] != 'Default':
                                    if policy['provider_filter']['name'] == invfilter['name']:
                                        ProvipSet = resolveFilter(rc, invfilter)
                                    if policy['consumer_filter']['name'] == invfilter['name']:
                                        ConsipSet = resolveFilter(rc, invfilter)
                        for a in ConsipSet:
                            for b in ProvipSet: 
                                if a != b:
                                    if rule['port_min'] == rule['port_max']:
                                        port = rule['port_min']
                                        print ("\t permit " + protocols[str(rule['proto'])]['Keyword'] + " host " + (a if policy['consumer_filter']['name'] != 'Default' else " any") + " host " + (b if policy['provider_filter']['name'] != 'Default' else " any") + " eq " + str(port))
                                        file1.write("\t permit " + protocols[str(rule['proto'])]['Keyword'] + " host " + (a if policy['consumer_filter']['name'] != 'Default' else " any") + " host " + (b if policy['provider_filter']['name'] != 'Default' else " any") + " eq " + str(port) + "\n")
                                    else:
                                        print ("\t permit " + protocols[str(rule['proto'])]['Keyword'] + " host " + (a if policy['consumer_filter']['name'] != 'Default' else " any") + " host " + (b if policy['provider_filter']['name'] != 'Default' else " any") + " range " + str(rule['port_min']) + "-" + str(rule['port_max']))
                                        file1.write("\t permit " + protocols[str(rule['proto'])]['Keyword'] + " host " + (a if policy['consumer_filter']['name'] != 'Default' else " any") + " host " + (b if policy['provider_filter']['name'] != 'Default' else " any") + " range " + str(rule['port_min']) + "-" + str(rule['port_max']) + "\n")
    print ("\t deny ip any any\n!\n\n")
    file1.write("\t deny ip any any\n!\n\n")
    file1.close()
    print (CYELLOW + "ACL Config File: ACL_config.txt created" + CEND)


# =================================================================================
# clean
# =================================================================================
def clean(restclient, root_scope_name):
    errors = []

    # Gather existing scopes and IDs
    resp = restclient.get('/openapi/v1/app_scopes/')
    if resp.status_code == 200:
        current_scopes = resp.json()
        root_scope = [
            x for x in current_scopes if x['name'] == root_scope_name]
        app_scope_id = root_scope[0]['id']
        vrf_id = root_scope[0]['query']['value']

    # -------------------------------------------------------------------------
    # DETERMINE SCOPES TO BE DELETED
    # Using two lists here as queues:
    # 1. toBeExamined is a FIFO where we add parent scopes at position zero and
    #    use pop to remove them from the end. We add one entire heirarchical
    #    level of parents before we add a single one of their children. This
    #    process will continue until there are no more children to add and the
    #    FIFO will eventually be empty.
    # 2. toBeDeleted is a LIFO where we append parent scopes at the end before
    #    we append their children. Later, we will pop scopes from the end when
    #    deleting them, so child scopes will always be deleted before their
    #    parents (which is required by Tetration).

    print ("[CHECKING] all scopes in Tetration.")
    toBeDeleted = []
    toBeExamined = [app_scope_id]
    while len(toBeExamined):
        scopeId = toBeExamined.pop()
        resp = restclient.get('/openapi/v1/app_scopes/' + scopeId)
        if resp.status_code == 200:
            for scope in resp.json()["child_app_scope_ids"]:
                toBeExamined.insert(0, scope)
                toBeDeleted.append(scope)
        else:
            print ("[ERROR] examining scope '{}'. This will cause problems deleting all scopes.".format(scopeId))
            errors.append("[ERROR] examining scope '{}'. This will cause problems deleting all scopes.".format(scopeId))
            print (resp, resp.text)

    # -------------------------------------------------------------------------
    # DELETE THE WORKSPACES
    # Walk through all applications and remove any in a scope that should be
    # deleted. In order to delete an application, we have to turn off enforcing
    # and make it secondary first.

    resp = restclient.get('/openapi/v1/applications/')
    if resp.status_code == 200:
        resp_data = resp.json()
    else:
        print ("[ERROR] reading application workspaces to determine which ones should be deleted.")
        errors.append("[ERROR] reading application workspaces to determine which ones should be deleted.")
        print (resp, resp.text)
        resp_data = {}
    for app in resp_data:
        appName = app["name"]
        if app["app_scope_id"] in toBeDeleted or app["app_scope_id"] == app_scope_id:
            app_id = app["id"]
            # first we turn off enforcement
            if app["enforcement_enabled"]:
                r = restclient.post('/openapi/v1/applications/' + app_id + '/disable_enforce')
                if r.status_code == 200:
                    print ("[CHANGED] app {} ({}) to not enforcing.".format(app_id, appName))
                else:
                    print ("[ERROR] changing app {} ({}) to not enforcing. Trying again...".format(app_id, appName))
                    time.sleep(1)
                    r = restclient.post('/openapi/v1/applications/' + app_id + '/disable_enforce')
                    if r.status_code == 200:
                        print ("[CHANGED] app {} ({}) to not enforcing.".format(app_id, appName))
                    else:
                        errors.append("[ERROR] Failed again. Details: {} -- {}".format(resp, resp.text))
                        print (resp, resp.text)
            # make the application secondary if it is primary
            if app["primary"]:
                req_payload = {"primary": "false"}
                r = restclient.put('/openapi/v1/applications/' + app_id, json_body=json.dumps(req_payload))
                if r.status_code == 200:
                    print ("[CHANGED] app {} ({}) to secondary.".format(app_id, appName))
                else:
                    # Wait and try again
                    print ("[ERROR] changing app {} ({}) to secondary. Trying again...".format(app_id, appName))
                    time.sleep(1)
                    r = restclient.post('/openapi/v1/applications/' + app_id + '/disable_enforce')
                    if r.status_code == 200:
                        print ("[CHANGED] app {} ({}) to not enforcing.".format(app_id, appName))
                    else:
                        errors.append("[ERROR] Failed again. Details: {} -- {}".format(resp, resp.text))
                        print (resp, resp.text)
            # now delete the app
            r = restclient.delete('/openapi/v1/applications/' + app_id)
            if r.status_code == 200:
                print ("[REMOVED] app {} ({}) successfully.".format(app_id, appName))
            else:
                # Wait and try again
                print ("[ERROR] deleting {} ({}). Trying again...".format(app_id, appName))
                time.sleep(1)
                r = restclient.delete('/openapi/v1/applications/' + app_id)
                if r.status_code == 200:
                    print ("[REMOVED] app {} ({}) successfully.".format(app_id, appName))
                else:
                    errors.append("[ERROR] Failed again. Details: {} -- {}".format(resp, resp.text))
                    print (resp, resp.text)

    # -------------------------------------------------------------------------
    # DETERMINE ALL FILTERS ASSOCIATED WITH THIS VRF_ID
    # Inventory filters have a query that the user enters but there is also a
    # query for the vrf_id to match. So we simply walk through all filters and
    # look for that query to match this vrf_id... if there is a match then
    # mark the filter as a target for deletion.  Before deleting filters,
    # we need to delete the agent config intents

    filtersToBeDeleted = []

    resp = restclient.get('/openapi/v1/filters/inventories')
    if resp.status_code == 200:
        resp_data = resp.json()
    else:
        print ("[ERROR] reading filters to determine which ones should be deleted.")
        errors.append("[ERROR] reading filters to determine which ones should be deleted.")
        print (resp, resp.text)
        resp_data = {}
    for filt in resp_data:
        try:
            inventory_filter_id = filt["id"]
            filterName = filt["name"]
            for query in filt["query"]["filters"]:
                if 'field' in query.iterkeys() and query["field"] == "vrf_id" and query["value"] == int(vrf_id):
                    filtersToBeDeleted.append({'id': inventory_filter_id, 'name': filterName})
        except:
            print(json.dumps(filt))

    # -------------------------------------------------------------------------
    # DELETE AGENT CONFIG INTENTS
    # Look through all agent config intents and delete instances that are based
    # on a filter or scope in filtersToBeDeleted or toBeDeleted (scopes)

    print ("[CHECKING] all inventory config intents in Tetration.")

    resp = restclient.get('/openapi/v1/inventory_config/intents')
    if resp.status_code == 200:
        resp_data = resp.json()
    else:
        print ("[ERROR] reading inventory config intents to determine which ones should be deleted.")
        errors.append("[ERROR] reading inventory config intents to determine which ones should be deleted.")
        print (resp, resp.text)
        resp_data = {}
    for intent in resp_data:
        intent_id = intent['id']
        filter_id = intent["inventory_filter_id"]
        if filter_id in filtersToBeDeleted or filter_id in toBeDeleted or filter_id == app_scope_id:
            r = restclient.delete('/openapi/v1/inventory_config/intents/' + intent_id)
            if r.status_code == 200:
                print ("[REMOVED] inventory config intent {}.".format(intent_id))
            else:
                print ("[ERROR] removing inventory config intent {}.".format(intent_id))
                errors.append("[ERROR] removing inventory config intent {}.".format(intent_id))
                print (r, r.text)

    # -------------------------------------------------------------------------
    # DELETE THE FILTERS

    while len(filtersToBeDeleted):
        filterId = filtersToBeDeleted.pop()
        r = restclient.delete('/openapi/v1/filters/inventories/' + filterId['id'])
        if r.status_code == 200:
            print ("[REMOVED] inventory filter {} named '{}'.".format(filterId['id'], filterId['name']))
        else:
            print ("[ERROR] removing inventory filter {} named '{}'.".format(filterId['id'], filterId['name']))
            errors.append("[ERROR] removing inventory filter {} named '{}'.".format(filterId['id'], filterId['name']))
            print (r, r.text)

    # -------------------------------------------------------------------------
    # DELETE THE SCOPES

    while len(toBeDeleted):
        scopeId = toBeDeleted.pop()
        resp = restclient.delete('/openapi/v1/app_scopes/' + scopeId)
        if resp.status_code == 200:
            print ("[REMOVED] scope {} successfully.".format(scopeId))
        else:
            print ("[ERROR] removing scope {}.".format(scopeId))
            errors.append("[ERROR] removing scope {}.".format(scopeId))
            print (resp, resp.text)


# =================================================================================
# Policies
# =================================================================================

def GetPolicies(rc, app_id):
    
    resp = rc.get('/applications/' + app_id + '/policies')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Policies list")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowPolicies(policies):
    """
        List all the Apps in Tetration Appliance
        Policy ID | Application ID | Version | Author | Scope ID | Primary | Enforced
        """
    data_list = []
    headers = ['Application ID', 'Name', 'Author', 'Scope ID', 'Primary', 'Enforced']
    for x in apps: data_list.append([x['id'],
                    x['name'], x['author'],
                    x['app_scope_id'], x['primary'], x['enforcement_enabled']]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def getDefaultDetail(rc, id):
    resp = rc.get('/applications/'+ id + '/default_policies')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Default Policies from your Apps"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json() 

def getAbsoluteDetail(rc, id):
    resp = rc.get('/applications/'+ id + '/absolute_policies')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Absolute Policies from your Apps"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json() 

def getCatchAllDetail(rc, id):
    resp = rc.get('/applications/'+ id + '/catch_all')
    if resp.status_code != 200:
        print(URED + "Failed to retrieve catch_all Policy from your Apps"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()


def GetDefaultInformation(default_policies):
    Default_Detail = []
    headers = ['Policy ID', 'Action', 'Provider', 'Consumer']
    try:
        for value in default_policies:
            Default_Detail.append([value["id"],value["action"],value["provider_filter"]["name"], value["consumer_filter"]["name"]])
        table = columnar(Default_Detail, headers, no_borders=False)
        print(table)
    except:
        print(CRED + "Default Information not found" +CEND)

def GetAbsoluteInformation(abs_policies):
    Absolute_Detail = []
    headers = ['Policy ID', 'Action', 'Provider', 'Consumer']
    try:
        for value in abs_policies:
            Absolute_Detail.append([value["id"],value["action"],value["provider_filter"]["name"], value["consumer_filter"]["name"]])
        table = columnar(Absolute_Detail, headers, no_borders=False)
        print(table)
    except:
        print(CRED + "Absolute Information not found" +CEND)


def GetServerPorts(rc):
    #get Server Ports config to root scope. Return server ports config. 
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: ")
    GetRootScope(vrfs)
    root_scope_name = input(CGREEN +"\nWhat is the root scope you want to check the server ports config: ")
    scopes = GetApplicationScopes(rc)
    root_scope_id = GetAppScopeId(scopes,root_scope_name)
    resp = rc.get("/openapi/v1/adm/" + root_scope_id + "/server_ports")
    if resp.status_code != 200:
        print(URED + "Failed to retrieve Server Ports Config list")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def uploadServerPorts(rc):
    #Upload Server Ports config to root scope
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: ")
    GetRootScope(vrfs)
    root_scope_name = input(CGREEN +"\nWhat is the root scope you want to upload server ports: ")
    scopes = GetApplicationScopes(rc)
    root_scope_id = GetAppScopeId(scopes,root_scope_name)
    file_path = "server_ports.txt"
    resp = rc.upload(file_path,'/adm/%s/server_ports' % root_scope_id,timeout=200)
    if resp.status_code == 200:
        print("\nUploaded sucessful!" + CEND)
    else:
        print("Error occured during upload server ports")
        print("Error code: "+str(resp.status_code))
        sys.exit(3)


def CreateCluster(rc):
    #Create new cluster under App. Return cluster_name and cluster_id
    Apps = GetApps(rc)
    AppsList = GetAppsNamewithID(Apps)
    print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: ")
    print(*AppsList, sep = "\n")
    app_name = input (CYELLOW + "\nWhich App name you want to add cluster in: ")
    app_id = GetAppsId(Apps, app_name)
    app_detail = getAppDetail(rc, app_id)
    app_versions = GetAppVersions(rc,app_id)
    version = GetLatestVersion(app_versions)
    print (CGREEN +"Here is the latest version of your app " +CYELLOW+ app_name + CGREEN +" : " + CYELLOW+version + CEND)
    cluster_name = input(CGREEN +"\nWhat is the name of the cluster under " + app_name + " app you want to create: ")
    query_type = input (CYELLOW + "\nHere are 2 types of query supported: Hostname or IP, which one you choose? ")
    if query_type == ("Hostname" or "hostname"): 
        hostname = input (CYELLOW + "\nWhat is the hostname contain? ")
        req_payload = {
        "name": cluster_name,
        "version": version,
        "description": "Created by API",
        "approved": True,
        "query": {
            "type": "contains",
            "field": "host_name",
            "value": hostname}
        }
    if query_type ==("IP" or "ip"): 
        ip_addr = input(CYELLOW + "\nWhat is the IP (Ex. 192.168.1.0/24 or 192.168.1.1): ")
        if ".0/" in ip_addr: 
            req_payload = {
            "name": cluster_name,
            "version": version,
            "description": "Created by API",
            "approved": True,
            "query": {
                "type": "subnet",
                "field": "ip",
                "value": ip_addr}
            }
        else:
            req_payload = {
            "name": cluster_name,
            "version": version,
            "description": "Created by API",
            "approved": True,
            "query": {
                "type": "eq",
                "field": "ip",
                "value": ip_addr}
                }

    else: print (CRED + "\nPlease key in the correct option for query (Hostname or IP) " + CEND)
    print("Adding cluster "+ CYELLOW+cluster_name + " into your app "+CYELLOW+app_name + CEND)
    resp = rc.post('/applications/' + app_id+ '/clusters', json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        cluster_id = str(parsed_resp["id"])
        print("\nCluster: "+ CYELLOW+cluster_name + " with ID " +CYELLOW+cluster_id+ " has been added into your app "+CYELLOW+app_name + CEND)
    else:
        print("Error occured during cluster creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)
    return cluster_name, cluster_id


def CatchAll(rc):
    #Change Catch All Action
    rc = CreateRestClient()
    Apps = GetApps(rc)
    AppsList = GetAppsNamewithID(Apps)

    print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: ")
    print(*AppsList, sep = "\n")
    app_name = input (CYELLOW + "\nWhich App name you want to change Catch All action: ")
    app_id = GetAppsId(Apps, app_name)

    app_versions = GetAppVersions(rc,app_id)
    version = GetLatestVersion(app_versions)
    print (CGREEN +"Here is the latest version of your app " +CYELLOW+ app_name + CGREEN +" : " + CYELLOW+version + CEND)
    catch_all = getCatchAllDetail(rc, app_id)
    print (CGREEN +"You App " +CYELLOW+ app_name + CGREEN +"'s current catch_all action is : " + CYELLOW+catch_all['action'] + CEND)
    choice = input (CYELLOW + "\nDo you want to change the action now (y/n)? ")
    if choice == "y" or choice == "Y": 
        if catch_all['action'] == "ALLOW": policy_action = "DENY" 
        else:  policy_action = "ALLOW"
        print(CGREEN +"Changing catch_all action for your app "+CYELLOW+app_name+ " now" +CEND)
        req_payload = {
        "version": version,
        "policy_action" : policy_action}
        resp = rc.put('/applications/' + app_id +'/catch_all', json_body=json.dumps(req_payload))
        parsed_resp = json.loads(resp.content)
        if resp.status_code == 200:
            print(Cyan + "\nCatch All Action has just been changed for your application " +CYELLOW+app_name+ CEND)
        else:
            print("Error occured during application creation")
            print("Error code: "+str(resp.status_code))
            print("Content: ")
            print(resp.content)
            sys.exit(3)
    else: sys.exit(3)


def CreatePolicy(rc):
    #Add policy without service port into application. Return: Policy_ID
    rc = CreateRestClient()
    prov_id = ""
    con_id = ""
    Apps = GetApps(rc)
    AppsList = GetAppsNamewithID(Apps)

    print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: ")
    print(*AppsList, sep = "\n")
    app_name = input (CYELLOW + "\nWhich App name you want to add policy in: ")
    app_id = GetAppsId(Apps, app_name)
    scope_id = GetScopeIdFromApps(Apps,app_name)

    app_versions = GetAppVersions(rc,app_id)
    version = GetLatestVersion(app_versions)
    print (CGREEN +"Here is the latest version of your app " +CYELLOW+ app_name + CGREEN +" : " + CYELLOW+version + CEND)
    policy_type = input (CYELLOW + "\n7. Default or absolute policy?: ")
    app_detail = getAppDetail(rc, app_id)
    with open('AppDetail.json', 'w') as outfile:
        json.dump(app_detail, outfile, indent=4)
    print (Cyan + "\nHere is the detail of your App " + app_name + " :")
    print (Cyan + json.dumps(app_detail, indent=4, sort_keys=True)+ CEND)
    
    provider_choice = input (CYELLOW + "\nWhich types of provider you want (Scope, Filter or Cluster): ")
    if provider_choice == "scope" or provider_choice == "Scope" or provider_choice == "s" or provider_choice == "S":
        scopes = GetApplicationScopes(rc)
        print (Cyan + "\nHere are the available scopes in your cluster: "+ CEND)
        ShowApplicationScopes(scopes)
        provider_scope =  input (CYELLOW + "\nWhich scope above (Root:Sub) you chooose?: ")
        prov_id = GetAppScopeId(scopes,name)
    if provider_choice == "Filter" or provider_choice == "filter" or provider_choice == "f" or provider_choice == "F":
        inventories = GetInventories(rc)
        print (CGREEN + "\nHere are the names and ID of all inventories in your cluster: ")
        GetInventoriesNamewithIDinScope(inventories,scope_id)
        provider_filter_name = input (CYELLOW + "\nWhich inventory name to be server (provider) in your policy: ")
        prov_id = GetInventoriesId(inventories, provider_filter_name)
    if provider_choice == "Cluster" or provider_choice == "cluster" or provider_choice == "c" or provider_choice == "C":
        clusters_list = GetClusters(rc, app_id)
        if clusters_list ==[]: 
            print (Cyan + "\nYour application " + app_name +" doesn't have any cluster. Please define it or start to try ADM now. " + CEND)
            print (CGREEN + "You can use different options (Scope, Filter) to create policies")
        else: 
            print (Cyan + "\nHere is the clusters (brief information)) of your App " + app_name + " :")
            ShowAppClusters(clusters_list)
            provider_cluster = input (CYELLOW + "\nWhich cluster name to be server (provider) in your policy: ")
            prov_id = GetClusterID(clusters_list, provider_cluster)

    consumer_choice = input (CYELLOW + "\nWhich types of consumer you want (Scope, Filter or Cluster): ")
    if consumer_choice == "scope" or consumer_choice == "Scope" or consumer_choice == "s" or consumer_choice == "S":
        scopes = GetApplicationScopes(rc)
        print (Cyan + "\nHere are the available scopes in your cluster: "+ CEND)
        ShowApplicationScopes(scopes)
        consumer_scope =  input (CYELLOW + "\nWhich scope above (Root:Sub) you chooose?: ")
        con_id = GetAppScopeId(scopes,consumer_scope)
    if consumer_choice == "Filter" or consumer_choice == "filter" or consumer_choice == "f" or consumer_choice == "F":
        inventories = GetInventories(rc)
        print (CGREEN + "\nHere are the names and ID of all inventories in your cluster: ")
        GetInventoriesNamewithIDinScope(inventories,scope_id)
        consumer_filter_name = input (CYELLOW + "\nWhich inventory name to be client (consumer) in your policy: ")
        con_id = GetInventoriesId(inventories, consumer_filter_name)
    if consumer_choice == "Cluster" or consumer_choice == "cluster" or consumer_choice == "c" or consumer_choice == "C":
        clusters_list = GetClusters(rc, app_id)
        if clusters_list ==[]: 
            print (Cyan + "\nYour application " + app_name +" doesn't have any cluster. Please define it or start to try ADM now. " + CEND)
            print (CGREEN + "You can use different options (Scope, Filter) to create policies")
        else: 
            print (Cyan + "\nHere is the clusters (brief information)) of your App " + app_name + " :")
            ShowAppClusters(clusters_list)
            consumer_cluster = input (CYELLOW + "\nWhich cluster name to be client (consumer) in your policy: ")
            con_id = GetClusterID(clusters_list, consumer_cluster)

    print ("Provider ID: " + prov_id)
    print ("consumer ID: " + con_id)
    
    policy_action = input (CYELLOW + "\nIs it Allow or deny?: ")

    print(CGREEN +"Adding Policy into your application "+CYELLOW+app_name+ CEND)
    req_payload = {
    "version": version,
    "rank" : policy_type,
    "policy_action" : policy_action,
    "priority" : 99,
    "consumer_filter_id" : con_id,
    "provider_filter_id" : prov_id}
    resp = rc.post('/applications/' + app_id +'/policies', json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        Policy_id = str(parsed_resp["id"])
        print(Cyan + "\nDefault Policy with ID " +CYELLOW+Policy_id + Cyan +" has just been added to your application " +CYELLOW+app_name+ CEND)
    else:
        print("Error occured during application creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)
    return Policy_id

def CreateAbsolutePolicy(rc):
    #Add Absolute policy without service port into application. Return: Policy_ID
    rc = CreateRestClient()
    prov_id = ""
    con_id = ""
    Apps = GetApps(rc)
    AppsList = GetAppsNamewithID(Apps)

    print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: ")
    print(*AppsList, sep = "\n")
    app_name = input (CYELLOW + "\nWhich App name you want to add Absolute policy in: ")
    app_id = GetAppsId(Apps, app_name)
    scope_id = GetScopeIdFromApps(Apps,app_name)

    app_versions = GetAppVersions(rc,app_id)
    version = GetLatestVersion(app_versions)
    print (CGREEN +"Here is the latest version of your app " +CYELLOW+ app_name + CGREEN +" : " + CYELLOW+version + CEND)
    app_detail = getAppDetail(rc, app_id)
    with open('AppDetail.json', 'w') as outfile:
        json.dump(app_detail, outfile, indent=4)
    print (Cyan + "\nHere is the detail of your App " + app_name + " :")
    print (Cyan + json.dumps(app_detail, indent=4, sort_keys=True)+ CEND)
    
    provider_choice = input (CYELLOW + "\nWhich types of provider you want (Scope, Filter or Cluster): ")
    if provider_choice == "scope" or provider_choice == "Scope" or provider_choice == "s" or provider_choice == "S":
        scopes = GetApplicationScopes(rc)
        print (Cyan + "\nHere are the available scopes in your cluster: "+ CEND)
        ShowApplicationScopes(scopes)
        provider_scope =  input (CYELLOW + "\nWhich scope above (Root:Sub) you chooose?: ")
        prov_id = GetAppScopeId(scopes,provider_scope)
    if provider_choice == "Filter" or provider_choice == "filter" or provider_choice == "f" or provider_choice == "F":
        inventories = GetInventories(rc)
        print (CGREEN + "\nHere are the names and ID of all inventories in your cluster: ")
        GetInventoriesNamewithIDinScope(inventories,scope_id)
        provider_filter_name = input (CYELLOW + "\nWhich inventory name to be server (provider) in your Absolute policy: ")
        prov_id = GetInventoriesId(inventories, provider_filter_name)
    if provider_choice == "Cluster" or provider_choice == "cluster" or provider_choice == "c" or provider_choice == "C":
        clusters_list = GetClusters(rc, app_id)
        if clusters_list ==[]: 
            print (Cyan + "\nYour application " + app_name +" doesn't have any cluster. Please define it or start to try ADM now. " + CEND)
            print (CGREEN + "You can use different options (Scope, Filter) to create policies")
        else: 
            print (Cyan + "\nHere is the clusters (brief information)) of your App " + app_name + " :")
            ShowAppClusters(clusters_list)
            provider_cluster = input (CYELLOW + "\nWhich cluster name to be server (provider) in your Absolute policy: ")
            prov_id = GetClusterID(clusters_list, provider_cluster)

    consumer_choice = input (CYELLOW + "\nWhich types of consumer you want (Scope, Filter or Cluster): ")
    if consumer_choice == "scope" or consumer_choice == "Scope" or consumer_choice == "s" or consumer_choice == "S":
        scopes = GetApplicationScopes(rc)
        print (Cyan + "\nHere are the available scopes in your cluster: "+ CEND)
        ShowApplicationScopes(scopes)
        consumer_scope =  input (CYELLOW + "\nWhich scope above (Root:Sub) you chooose?: ")
        con_id = GetAppScopeId(scopes,consumer_scope)
    if consumer_choice == "Filter" or consumer_choice == "filter" or consumer_choice == "f" or consumer_choice == "F":
        inventories = GetInventories(rc)
        print (CGREEN + "\nHere are the names and ID of all inventories in your cluster: ")
        GetInventoriesNamewithIDinScope(inventories,scope_id)
        consumer_filter_name = input (CYELLOW + "\nWhich inventory name to be client (consumer) in your Absolute policy: ")
        con_id = GetInventoriesId(inventories, consumer_filter_name)
    if consumer_choice == "Cluster" or consumer_choice == "cluster" or consumer_choice == "c" or consumer_choice == "C":
        clusters_list = GetClusters(rc, app_id)
        if clusters_list ==[]: 
            print (Cyan + "\nYour application " + app_name +" doesn't have any cluster. Please define it or start to try ADM now. " + CEND)
            print (CGREEN + "You can use different options (Scope, Filter) to create policies")
        else: 
            print (Cyan + "\nHere is the clusters (brief information)) of your App " + app_name + " :")
            ShowAppClusters(clusters_list)
            consumer_cluster = input (CYELLOW + "\nWhich cluster name to be client (consumer) in your Absolute policy: ")
            con_id = GetClusterID(clusters_list, consumer_cluster)

    print ("Provider ID: " + prov_id)
    print ("consumer ID: " + con_id)
    
    policy_action = input (CYELLOW + "\nIs it Allow or deny?: ")

    print(CGREEN +"Adding Absolute Policy into your application "+CYELLOW+app_name+ CEND)
    req_payload = {
    "version": version,
    "policy_action" : policy_action,
    "priority" : 99,
    "consumer_filter_id" : con_id,
    "provider_filter_id" : prov_id}
    resp = rc.post('/applications/' + app_id +'/absolute_policies', json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        Policy_id = str(parsed_resp["id"])
        print(Cyan + "\nAbsolute Policy with ID " +CYELLOW+Policy_id + Cyan +" has just been added to your application " +CYELLOW+app_name+ CEND)
    else:
        print("Error occured during application creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)
    return Policy_id

def CreateDefaultPolicy(rc):
    #Add default policy without service port into application. Return: Policy_ID
    rc = CreateRestClient()
    prov_id = ""
    con_id = ""
    Apps = GetApps(rc)
    AppsList = GetAppsNamewithID(Apps)

    print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: ")
    print(*AppsList, sep = "\n")
    app_name = input (CYELLOW + "\nWhich App name you want to add Default policy in: ")
    app_id = GetAppsId(Apps, app_name)
    scope_id = GetScopeIdFromApps(Apps,app_name)

    app_versions = GetAppVersions(rc,app_id)
    version = GetLatestVersion(app_versions)
    print (CGREEN +"Here is the latest version of your app " +CYELLOW+ app_name + CGREEN +" : " + CYELLOW+version + CEND)
    app_detail = getAppDetail(rc, app_id)
    with open('AppDetail.json', 'w') as outfile:
        json.dump(app_detail, outfile, indent=4)
    print (Cyan + "\nHere is the detail of your App " + app_name + " :")
    print (Cyan + json.dumps(app_detail, indent=4, sort_keys=True)+ CEND)
    
    provider_choice = input (CYELLOW + "\nWhich types of provider you want (Scope, Filter or Cluster): ")
    if provider_choice == "scope" or provider_choice == "Scope" or provider_choice == "s" or provider_choice == "S":
        scopes = GetApplicationScopes(rc)
        print (Cyan + "\nHere are the available scopes in your cluster: "+ CEND)
        ShowApplicationScopes(scopes)
        provider_scope =  input (CYELLOW + "\nWhich scope above (Root:Sub) you chooose?: ")
        prov_id = GetAppScopeId(scopes,name)
    if provider_choice == "Filter" or provider_choice == "filter" or provider_choice == "f" or provider_choice == "F":
        inventories = GetInventories(rc)
        print (CGREEN + "\nHere are the names and ID of all inventories in your cluster: ")
        GetInventoriesNamewithIDinScope(inventories,scope_id)
        provider_filter_name = input (CYELLOW + "\nWhich inventory name to be server (provider) in your Default policy: ")
        prov_id = GetInventoriesId(inventories, provider_filter_name)
    if provider_choice == "Cluster" or provider_choice == "cluster" or provider_choice == "c" or provider_choice == "C":
        clusters_list = GetClusters(rc, app_id)
        if clusters_list ==[]: 
            print (Cyan + "\nYour application " + app_name +" doesn't have any cluster. Please define it or start to try ADM now. " + CEND)
            print (CGREEN + "You can use different options (Scope, Filter) to create policies")
        else: 
            print (Cyan + "\nHere is the clusters (brief information)) of your App " + app_name + " :")
            ShowAppClusters(clusters_list)
            provider_cluster = input (CYELLOW + "\nWhich cluster name to be server (provider) in your Default policy: ")
            prov_id = GetClusterID(clusters_list, provider_cluster)

    consumer_choice = input (CYELLOW + "\nWhich types of consumer you want (Scope, Filter or Cluster): ")
    if consumer_choice == "scope" or consumer_choice == "Scope" or consumer_choice == "s" or consumer_choice == "S":
        scopes = GetApplicationScopes(rc)
        print (Cyan + "\nHere are the available scopes in your cluster: "+ CEND)
        ShowApplicationScopes(scopes)
        consumer_scope =  input (CYELLOW + "\nWhich scope above (Root:Sub) you chooose?: ")
        con_id = GetAppScopeId(scopes,consumer_scope)
    if consumer_choice == "Filter" or consumer_choice == "filter" or consumer_choice == "f" or consumer_choice == "F":
        inventories = GetInventories(rc)
        print (CGREEN + "\nHere are the names and ID of all inventories in your cluster: ")
        GetInventoriesNamewithIDinScope(inventories,scope_id)
        consumer_filter_name = input (CYELLOW + "\nWhich inventory name to be client (consumer) in your Default policy: ")
        con_id = GetInventoriesId(inventories, consumer_filter_name)
    if consumer_choice == "Cluster" or consumer_choice == "cluster" or consumer_choice == "c" or consumer_choice == "C":
        clusters_list = GetClusters(rc, app_id)
        if clusters_list ==[]: 
            print (Cyan + "\nYour application " + app_name +" doesn't have any cluster. Please define it or start to try ADM now. " + CEND)
            print (CGREEN + "You can use different options (Scope, Filter) to create policies")
        else: 
            print (Cyan + "\nHere is the clusters (brief information)) of your App " + app_name + " :")
            ShowAppClusters(clusters_list)
            consumer_cluster = input (CYELLOW + "\nWhich cluster name to be client (consumer) in your Default policy: ")
            con_id = GetClusterID(clusters_list, consumer_cluster)

    print ("Provider ID: " + prov_id)
    print ("consumer ID: " + con_id)
    
    policy_action = input (CYELLOW + "\nIs it Allow or deny?: ")

    print(CGREEN +"Adding Default Policy into your application "+CYELLOW+app_name+ CEND)
    req_payload = {
    "version": version,
    "policy_action" : policy_action,
    "priority" : 99,
    "consumer_filter_id" : con_id,
    "provider_filter_id" : prov_id}
    resp = rc.post('/applications/' + app_id +'/default_policies', json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        Policy_id = str(parsed_resp["id"])
        print(Cyan + "\nDefault Policy with ID " +CYELLOW+Policy_id + Cyan +" has just been added to your application " +CYELLOW+app_name+ CEND)
    else:
        print("Error occured during application creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)
    return Policy_id


def AddServicePort(rc):
    #Add policy without service port into application. Return: L4_Param_ID
    rc = CreateRestClient()
    Apps = GetApps(rc)
    AppsList = GetAppsNamewithID(Apps)

    print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: ")
    print(*AppsList, sep = "\n")

    app_name = input (CYELLOW + "\nWhich App name you want to add policy in: ")
    app_id = GetAppsId(Apps, app_name)
    app_versions = GetAppVersions(rc,app_id)
    version = GetLatestVersion(app_versions)
    print (CGREEN +"Here is the latest version of your app " +CYELLOW+ app_name + CGREEN +" : " + CYELLOW+version + CEND)
    app_detail = getAppDetail(rc, app_id)
    with open('AppDetail.json', 'w') as outfile:
        json.dump(app_detail, outfile, indent=4)
    print (Cyan + "\nHere is the detail of your App " + app_name + " :")
    print (Cyan + json.dumps(app_detail, indent=4, sort_keys=True)+ CEND)
    policy_type = input (CYELLOW + "\nWhich policy type (default or absolute) you want to add service port? ")
    if policy_type == "default" or policy_type =="Default" or policy_type =="Def" or policy_type == "def":
        Default_Policies = getDefaultDetail(rc,app_id)
        print (CGREEN +"Here are all default policies defined in your apps " +CYELLOW+ app_name + CEND)
        GetDefaultInformation(Default_Policies)
    if policy_type == "absolute" or policy_type == "Absolute" or policy_type == "abs" or policy_type == "Abs":
        Absolute_Policies = getAbsoluteDetail(rc,app_id)
        print (CGREEN +"Here are all Absolute policies defined in your apps " +CYELLOW+ app_name + CEND)
        GetAbsoluteInformation(Absolute_Policies)

    pol_id = input (CYELLOW + "\nWhich policy ID above (copy - paste) you want to add service port? "+ CEND)
    port_choice = input (CYELLOW + "\nTCP, UDP, ICMP or Any? "+ CEND)
    if port_choice == ("TCP" or "tcp"): proto = 6
    if port_choice == ("UDP" or "udp"): proto = 17
    if port_choice == ("ICMP" or "icmp"): proto = 1
    if port_choice == ("Any" or "ANY" or "any"): proto = null
    start_port = input (CYELLOW + "\nFrom which port? "+ CEND)
    end_port = input (CYELLOW + "\nTo which port? "+ CEND)
    print(CGREEN +"Adding Service Ports into your policy ID "+CYELLOW+pol_id+ CEND)
    req_payload = {
    "version": version,
    "start_port" : start_port,
    "end_port" : end_port,
    "proto" : proto,
    "description" : "Updated by API"
    }
    resp = rc.post('/policies/' + pol_id + '/l4_params', json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        L4_id = str(parsed_resp["l4_params"][0]["id"])
        print(Cyan + "\nA service port with " +CYELLOW+L4_id + Cyan +" ID has just been added to your application " +CYELLOW+app_name+ CEND)
    else:
        print("Error occured during application creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)
    return L4_id

# =================================================================================
# Applications
# =================================================================================
def GetApps(rc):
    resp = rc.get('/applications')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Apps list")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowApps(Apps):
    AppsList = []
    headers = ['Number', 'App Name', 'Author', 'App ID', 'Primary?']
    for i,app in enumerate(Apps): AppsList.append([i+1,app["name"] , app['author'], app["id"], app['primary']])
    table = columnar(AppsList, headers, no_borders=False)
    print(table)

def GetAppsId(Apps, name):
    try:
        for app in Apps: 
            if name == app["name"]: return app["id"]
    except:
        print(URED + "Failed to retrieve App ID "+ CEND)

def GetScopeIdFromApps(Apps, name):
    for app in Apps: 
        if name == app["name"]: 
            return app["app_scope_id"]


def GetLatestVersion(app_versions):
    try:
        for vers in app_versions: 
            if "v" in vers["version"]: return vers["version"]
    except:
        print(URED + "Failed to retrieve latest app version"+ CEND)


def GetAppsName(Apps):
    AppsNames = []
    try:
        for app in Apps: 
            AppsNames.append(app["name"])
        return AppsNames
    except:
        print(URED + "Failed to retrieve Apps Names list"+ CEND)  

def GetAppsNamewithID(Apps):
    AppsList = []
    try:
        for app in Apps: 
            AppsList.append([app["name"] , app["id"]])
        return AppsList
    except:
        print(URED + "Failed to retrieve Apps names with ID list"+ CEND) 

def getAppDetail(rc, id):
    resp = rc.get('/applications/'+ id)

    if resp.status_code != 200:
        print(URED + "Failed to retrieve App detail"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json() 

def getEnforcedApps(Apps):
    EnforcedList = []
    for app in Apps:
        if app["enforcement_enabled"] == True:
           EnforcedList.append([app["name"] , "Enforced"])
    return EnforcedList

def ShowAppDetail(details):
    """
        List application detail in Tetration Appliance
        Application ID | App Name | Author | Scope ID | Primary | Enforced
        """
    data_list = []
    headers = ['Application ID', 'Name', 'Author', 'Scope ID', 'Primary', 'Enforced']
    data_list.append([details['id'],
                    details['name'], details['author'],
                    details['app_scope_id'], details['primary'], details['enforcement_enabled']]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)


def GetAppInfor(rc, appid):
    resp = rc.get('/applications/' + appid + '/details')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Apps list" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowAppInfor(details):
    """
        List application detail in Tetration Appliance
        Application ID | App Name | Author | Scope Name | Primary | Enforced | Version | Policy Priority | Catch All Action
        """
    data_list = []
    headers = ['Application ID', 'Name', 'Author', 'Scope Name', 'Primary', 'Enforced', 'Version', 'Policy Priority', 'Catch All Action']
    data_list.append([details['id'],
                    details['name'], details['author'],
                    details['app_scope']['name'], details['primary'], details['enforcement_enabled'], details['version'], details['app_scope']['policy_priority'], details['catch_all_action']]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetAppVersions(rc, appid):
    resp = rc.get('/applications/' + appid + '/versions')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Apps list" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowAppVersions(versions):
    """
        List all the Apps in Tetration Appliance
        Application Version | Created At | Name | Description
        """
    data_list = []
    headers = ['Application Version', 'Created At', 'Name', 'Description']
    for x in versions: data_list.append([x['version'], time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(x['created_at'])), 
                    x['name'], x['description']]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetClusters(rc, appid):
    resp = rc.get('/applications/' + appid + '/clusters')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Clusters list" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetClustersName(clusters):
    Clusters_Detail = []
    try:
        for value in clusters:
            Clusters_Detail.append([value["id"],value["name"],value["approved"]])
        return Clusters_Detail
    except:
        print(CRED + "Clusters detail not found" +CEND)

def GetClusterID(clusters, name):
    try:
        for value in clusters:
            if name == value["name"]: return value["id"]
    except:
        print(CRED + "Cluster ID not found" +CEND)

def ShowAppClusters(clusters):
    """
        List all the clusters in an App in Tetration Appliance
        Cluster ID | Name | Approved
        """
    data_list = []
    headers = ['Cluster ID', 'Name', 'Approved']
    for value in clusters:
        data_list.append([value["id"],value["name"],value["approved"]])
    table = columnar(data_list, headers, no_borders=False)
    print(table)
    
def CreateApp(rc, scopes, scope):
    """Create Apps Workspace under Scope without policy, if you want to add policies, use function in Policies folder
    Returns:
        Apps ID, Apps Workspace Name 
    """
    apps_name = input(CGREEN +"\nWhat is the name of your apps under " + scope + " scope you want to create: ")
    app_scope_id = GetAppScopeId(scopes,scope)
    catch_all_action = input(CGREEN +"\nBlackList or Whitelist apps (ALLOW or DENY by default): ")
    print("Building Application: "+CYELLOW+apps_name+ " under Scope " +CYELLOW+scope+ " without policy for you" + CEND)
    req_payload = {
    "name": apps_name,
    "app_scope_id": app_scope_id,
    "description": "Created by Tetration API",
    "primary": False,
    "alternate_query_mode": True,
    "enforcement_enabled": False,
    "absolute_policies": [],
    "default_policies": [],
    "catch_all_action": catch_all_action}
    resp = rc.post('/applications', json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        apps_id = str(parsed_resp["id"])
        print("\nApplication: "+CYELLOW+apps_name+ " with ID " +CYELLOW+apps_id +" has been created to stop attacks from Attacker to your Struts container" + CEND)
    else:
        print("Error occured during application creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)
    return apps_name, apps_id


# =================================================================================
# Security
# =================================================================================

def downloadHash(rc):
    #Download whitelist or blacklist hash to Tetration root Scope. Sample csv: HashType,FileHash,FileName,Notes (SHA‐1,1AF17E73721DBE0C40011B82ED4BB1A7DBE3CE29,application_1.exe,Sample Notes)
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
    GetRootScope(vrfs)
    root_scope_name = input(CGREEN +"\nWhat is the root scope you want to download filehash: "+ CEND)
    b_w_List = input(CGREEN +"\nIs it blacklist or whitelist: ")
    resp = rc.download("FileHashDown.csv", "/assets/user_filehash/download/" +root_scope_name+ "/" + b_w_List)
    if resp.status_code == 200:
        print("\nDownloaded sucessful!" + CEND)
    else:
        print("Error occured during download hash")
        print("Error code: "+str(resp.status_code))
        sys.exit(3)


def uploadHash(rc):
    #Upload whitelist or blacklist hash to Tetration root Scope. Sample csv: HashType,FileHash,FileName,Notes (SHA‐1,1AF17E73721DBE0C40011B82ED4BB1A7DBE3CE29,application_1.exe,Sample Notes)
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
    GetRootScope(vrfs)
    root_scope_name = input(CGREEN +"\nWhat is the root scope you want to upload filehash: "+ CEND)
    b_w_List = input(CGREEN +"\nIs it blacklist or whitelist: "+ CEND)
    file_path = "sampleFileHashUpload.csv"
    resp = rc.upload(file_path, "/assets/user_filehash/upload/" + root_scope_name + "/" + b_w_List)
    if resp.status_code == 200:
        print("\nUploaded sucessful!" + CEND)
    else:
        print("Error occured during upload hash")
        print("Error code: "+str(resp.status_code))
        sys.exit(3)


def deleteHash(rc):
    #Delete whitelist or blacklist hash in Tetration root Scope.
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
    GetRootScope(vrfs)
    root_scope_name = input(CGREEN +"\nWhat is the root scope you want to upload filehash: "+ CEND)
    b_w_List = input(CGREEN +"\nIs it blacklist or whitelist: ")
    file_path = "FileHashDelete.csv"
    resp = rc.upload(file_path, "/assets/user_filehash/delete/" + root_scope_name + "/" + b_w_List)
    if resp.status_code == 200:
        print("\nDeleted sucessful!" + CEND)
    else:
        print("Error occured during delete hash")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)


# =================================================================================
# Roles and Users
# =================================================================================
def getRoles(rc):
    resp = rc.get('/roles')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve show inv")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowRoles(roles):
    """
        List all the Roles in Tetration Appliance
        Role ID | Role Name | Description
        """
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Role ID', 'Name', 'Description']
        data_list = [[x['id'],
                    x['name'], x['description']] for x in roles ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)


def GetRoleId(roles, name):
    try:
        for role in roles:
            if name == role["name"]:
                print ("Here is your Role ID: " + role["id"])
                return role["id"]
            else: continue
    except:
        print(URED + "Role {name} not found".format(name=name))

def getRoleDetail(rc, id):
    resp = rc.get('/roles/'+ id)

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Role detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json() 

def GetRolesNamewithID(roles):
    RolesList = []
    headers = ['Role Name', 'Role ID']
    for role in roles:
        RolesList.append([role["name"] , role["id"]])
    table = columnar(RolesList, headers, no_borders=False)
    print(table) 

def ShowRoleDetails(roles):
    """
        List all the Roles in Tetration Appliance
        Role ID | Role Name | Description | Ability
        """
    headers = ['Role ID', 'Name', 'Description']
    data_list = []
    data_list.append ([roles['id'],
                    roles['name'], roles['description']])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def getUsers(rc):
    resp = rc.get('/users')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve show inv")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetRoleID(roles, name):
    try:
        return [role["id"] for role in roles if role["name"] == name][0]
    except:
        print(CRED + "The Role {name} is not exist. Try again".format(name=name))


def ShowUsers(users, roles):
    """
        List all the Users in Tetration Appliance
        User ID | Name | Email | Roles
        """
    columns = None
    dict_roles = {r['id']: r['name'] for r in roles}
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['User ID', 'Name', 'e-Mail', 'Roles']
        data_list = [[x['id'],
                    '{0} {1}'.format(x['first_name'],x['last_name']),
                    x['email'],
                    ','.join([dict_roles[y] for y in x['role_ids']])] for x in users ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetUsersEmailwithID(Users):
    UsersList = []
    try:
        for User in Users: 
            UsersList.append([User["email"] , User["id"]])
        return UsersList
    except:
        print(CRED + "Failed to retrieve Users list") 

def GetUserId(Users, email):
    try:
        for User in Users:
            if email == User["email"]:
                print ("Here is your User ID: " + User["id"])
                return User["id"]
            else: continue
    except:
        print(CRED + "User {email} not found".format(email=email))

def getUserDetail(rc, id):
    resp = rc.get('/users/'+ id)

    if resp.status_code != 200:
        print(CRED + "Failed to retrieve User detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()  

def ShowUserDetail(users):
    """
        List details of a user in Tetration Appliance
        User ID | Email | First Name | Last Name | Root Scope ID | Last signed in
        """
    data_list = []
    headers = ['User ID', 'Email', 'First Name', 'Last Name', 'Root Scope ID', 'Last signed in']
    data_list.append([users['id'],
                    users['email'], users['first_name'],
                    users['last_name'], users['preferences']['root_app_scope_id'], time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(users['current_sign_in_at']))]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def CreateRole(rc):
    name = input (CGREEN + "\nKey in the name of the Role you want to create: "+ CEND)
    Roles = getRoles (rc)
    for role in Roles:
        if name == role["name"]:
            print(URED + "\nRole {name} is conflict with existing Role filter. Please choose different name".format(name=name)+ CEND)
        else:
            scopes = GetApplicationScopes(rc)
            vrfs = GetVRFs(rc)
            print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
            GetRootScope(vrfs)
            scope = input (CGREEN + "\nWhat is the Root scope you want your filter belong to: "+ CEND)
            scope_id = GetAppScopeId(scopes,scope)
            print(CGREEN + "\nBuilding Role: "+CYELLOW+name+ CGREEN + " under Scope " +CYELLOW+scope+ CEND)
            req_payload = {
            "app_scope_id": scope_id,
            "description": "Created by Python",
            "name": name
            }
            resp = rc.post("/roles", json_body=json.dumps(req_payload))
            parsed_resp = json.loads(resp.content)
            if resp.status_code == 200:
                Role_id = parsed_resp["id"]
                print(Cyan + "\nRole: "+CYELLOW+name+ Cyan + " with Role ID: " + CYELLOW + Role_id + Cyan + " has been created"+ CEND)
            else:
                print("Error occured during sub scope creation")
                print("Error code: "+str(resp.status_code))
                print("Content: ")
                print(resp.content)
                sys.exit(3)
            return name, Role_id

def CreateUser(rc):
    email = input (CGREEN + "\nKey in the email of the User you want to create: "+ CEND)
    Users = getUsers (rc)
    for User in Users:
        if email == User["email"]:
            print(CRED + "\nUser with {email} is conflict with existing User filter. Please choose different email".format(email=email)+ CEND)
        else:
            scopes = GetApplicationScopes(rc)
            scope_name = input (CGREEN + "\nWhat is the scope (Root:Sub) you want your user belong to: "+ CEND)
            scope_id = GetAppScopeId(scopes,scope_name)
            first_name = input (CGREEN + "\nWhat is the firstname of your user: "+ CEND)
            last_name = input (CGREEN + "\nWhat is the lastname of your user: "+ CEND)
            Roles = getRoles(rc)
            print (CGREEN + "\nHere are the names and ID of all Roles in your cluster: "+ CEND)
            GetRolesNamewithID(Roles)
            role_name = input (CGREEN + "\nWhich role above you want your user has: "+ CEND)
            role_id = GetRoleId(Roles, role_name)
            print(CGREEN + "\nCreating User: "+CYELLOW+first_name+ CGREEN + " under Scope " +CYELLOW+scope_name+ CGREEN + " and Role " + CYELLOW+role_name+ CEND)
            req_payload = {
            "first_name": first_name,
            "last_name": last_name,
            "email": email,
            "app_scope_id": scope_id,
            "role_ids": role_id
            }
            resp = rc.post("/users", json_body=json.dumps(req_payload))
            parsed_resp = json.loads(resp.content)
            if resp.status_code == 200:
                User_id = parsed_resp["id"]
                print(Cyan + "\nUser: "+CYELLOW+first_name+ Cyan + " with User ID: " + CYELLOW + User_id + Cyan + " has been created"+ CEND)
            else:
                print("Error occured during user creation")
                print("Error code: "+str(resp.status_code))
                print("Content: ")
                print(resp.content)
                sys.exit(3)
            return email, User_id

def role2Scope(rc):
    roles = getRoles(rc)
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "\nHere is the list of Roles in Tetration cluster: " + CEND)
    ShowRoles(roles)
    role_name = input (CGREEN + "\nKey in the name of the Role you want to assign to scope: "+ CEND)
    role_id = GetRoleID (roles, role_name)
    scope_name = input (CGREEN + "\nWhich scope (Root:Sub) you want to assign the Role to: "+ CEND)
    scope_id = GetAppScopeId (scopes,scope_name)
    ability = input (CGREEN + "\nWhich ability (SCOPE_READ, SCOPE_WRITE, EXECUTE, ENFORCE, SCOPE_OWNER, DEVELOPER) you want to assign for this role: "+ CEND)
    print(CGREEN + "\nAssigning Role: "+CYELLOW+role_name+ CGREEN + " into Scope " +CYELLOW+scope_name+ CEND)
    req_payload = {
    "app_scope_id": scope_id,
    "ability": ability
    }
    resp = rc.post("/roles/" + role_id + "/capabilities", json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        print(Cyan + "\nRole: "+CYELLOW+role_name+ Cyan + " with " +CYELLOW+ability+ Cyan+ " assigned to " + CYELLOW + scope_name + CEND)
    else:
        print("Error occured during assigning role to scope")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)
    return role_id, scope_id


# =================================================================================
# Flows
# =================================================================================
def GetFlowDimensions(rc):
    resp = rc.get('/flowsearch/dimensions')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve dimensions list"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetFlowMetrics(rc):
    resp = rc.get('/flowsearch/metrics')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve metrics list"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

# =================================================================================
# Orchestrators
# =================================================================================
def GetOrchestrators(rc):
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: " + CEND)
    GetRootScope(vrfs)
    root_scope_name = input(CGREEN +"\nWhat is the root scope you want to get the orchestrators: "+ CEND)
    resp = rc.get('/orchestrator/' + root_scope_name)

    if resp.status_code != 200:
        print(URED + "Failed to retrieve show orc"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetOrchestratorsScope(rc,scope):
    resp = rc.get('/orchestrator/' + scope)

    if resp.status_code != 200:
        print(URED + "Failed to retrieve show orc" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowOrchestrators(orchestrators):
    """
        List all the Apps in Tetration Appliance
        Orchestrator ID | Created At | Updated At | Name | Type | Host | Port | InSecure
        """
    data_list = []
    headers = ['Orchestrator ID', 'Created At', 'Updated At', 'Name', 'Type', 'Host', 'Port', 'InSecure']
    for x in orchestrators: 
        data_list.append([x['id'], time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(x['created_at'])), time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(x['updated_at'])), x['name'], x['type'], x['hosts_list'][0]['host_name'], x['hosts_list'][0]['port_number'], x['insecure']]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetOrchestratorsNamewithID(Orchestrators):
    OrchestratorsList = []
    headers = ['Orchestrator name', 'ID', 'Type']
    try:
        for orc in Orchestrators: 
            OrchestratorsList.append([orc["name"] , orc["id"] , orc["type"]])
        table = columnar(OrchestratorsList, headers, no_borders=False)
        print(table)
    except:
        print(URED + "Failed to retrieve show orc" + CEND) 

def GetOrchestratorsId(Orchestrators, name):
    try:
        for orc in Orchestrators:
            if name == orc["name"]:
                print ("Here is your Orchestrator ID: " + orc["id"]+ CEND)
                return orc["id"]
            else: continue
    except:
        print(URED + "Orchestrator {name} not found".format(name=name) + CEND)

def getOrchestratorDetail(rc, scope, orch_id):
    resp = rc.get('/orchestrator/'+ scope + '/' + orch_id)

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Orchestrator detail" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowOrchestrator(orchestrators):
    """
        List all the Apps in Tetration Appliance
        Orchestrator ID | Created At | Updated At | Name | Type | Host | Port | InSecure
        """
    data_list = []
    headers = ['Orchestrator ID', 'Created At', 'Updated At', 'Name', 'Type', 'Host', 'Port', 'InSecure']
    data_list.append([orchestrators['id'], time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(orchestrators['created_at'])), time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(orchestrators['updated_at'])), orchestrators['name'], orchestrators['type'], orchestrators['hosts_list'][0]['host_name'], orchestrators['hosts_list'][0]['port_number'], orchestrators['insecure']]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def CreateK8sOrchestrator(rc, scope):
    """Create K8s Orchestrator under a Scope
    Returns:
        Orchestrator ID and Name 
    """
    orch_name = input("\nName of the K8s orchestrator under " + scope + " scope you want to create: "+ CEND)
    address = input("\nThe IP address for master Node: ")
    port = int(input("\nThe port for Master Node: "))
    print("Building orchestrator: "+CYELLOW+orch_name+ " under Scope " +CYELLOW+scope + CEND)
    
    # Now build the orchestrator, username and password is mandatory, although they can be black
    req_payload = {
    "name": orch_name,
    "type": "kubernetes",
    "hosts_list": [ 
        { "host_name": address, 
        "port_number": port}],
        "username":"",
        "password":"",
        "ca_certificate":"-----BEGIN CERTIFICATE-----\nMIIC6jCCAdKgAwIBAgIBATANBgkqhkiG9w0BAQsFADAmMSQwIgYDVQQDDBtvcGVu\nc2hpZnQtc2lnbmVyQDE1ODA5NTkyODUwHhcNMjAwMjA2MDMyMTI0WhcNMjUwMjA0\nMDMyMTI1WjAmMSQwIgYDVQQDDBtvcGVuc2hpZnQtc2lnbmVyQDE1ODA5NTkyODUw\nggEiMA0GCSqGSIb3DQEBAQUAA4IBDwAwggEKAoIBAQDRUMiffbnwVX8wIhbG2RJa\nZ9FwbEWTRBYPL40J15JxsNVuwCOicH5jDT/fWS4IKLui45JKgckVmajxjg9YmXUA\nFP3srYBSrdDaFJL8cFgb0HmCSxyi6GYBRmKeEwwIk3QlvBcxXAmWEyBeTWvfuzkF\nr/82jDJ9UUi5lc6U/risKA8NlV+P6OAmzHcwDYpMXpGXgtSQDjlIHMJ/oCTVbUOz\nWDpVplw3+KALDM5r4s3YOevvGlDZCb3FrqFEZz5qdD3cDUhOADX3dlgrFe/QhZhE\nyhNxtHgy7UT0IXrcq4BIsnOzMUNhhjrHUNIy3fk17YMgQCOyQGwlP7zE1zVvAGDv\nAgMBAAGjIzAhMA4GA1UdDwEB/wQEAwICpDAPBgNVHRMBAf8EBTADAQH/MA0GCSqG\nSIb3DQEBCwUAA4IBAQBb3NZml7hmI1ix5uXzHzPpPTwddKOxRfV1Clwi+wzKXlOU\nJHPXigE+AaPHwS9BJQNEmR0UIWChg4xn7ff9y14QI2GcM77KPJsORvoC5TgGKJtE\nS1pH/rodufqT7+J/lF9Uh78BYzmoaFRbsbHNsjmd7oXZ2rF1C8O9xCMg8TGLDXQP\nZW3mRp9vDjOtIN/DOjxG1VqljgqXwNwCrCumqMPyLwqmgP0/QzQcyEhgI+Tw2505\n2LqgYI9/2c5BfTL2vb5lgvxorYfGOQib7QClow/2y3vmHJ/9EPlpKNX6L7/BpYVx\nmEPsHa63aUaDCjQwKv7hHO80rU1AJisRQkegiJHt\n-----END CERTIFICATE-----",
        "certificate": "-----BEGIN CERTIFICATE-----\nMIIDJDCCAgygAwIBAgIBBjANBgkqhkiG9w0BAQsFADAmMSQwIgYDVQQDDBtvcGVu\nc2hpZnQtc2lnbmVyQDE1ODA5NTkyODUwHhcNMjAwMjA2MDMyMTI2WhcNMjIwMjA1\nMDMyMTI3WjBOMTUwFQYDVQQKEw5zeXN0ZW06bWFzdGVyczAcBgNVBAoTFXN5c3Rl\nbTpjbHVzdGVyLWFkbWluczEVMBMGA1UEAxMMc3lzdGVtOmFkbWluMIIBIjANBgkq\nhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEA6XbI5LgmDK/Ftc5NvppOL14e7fUHoyCa\nmWZYPPSSBpsEAMEbK2N6npaWURsLxX/soh6MDteMYal53YDx9YS+Yks+JMIXE8tx\nCdDBCYRE62JFcX/1+VIJoAU3DcREK71T9r4Iz0PDAiw7PGvIJBNRe6MZasSsxrZp\nJp8ZiYAGjKQLEU/cdevXZDxIQfh/7kPeYyy+ANsYifCRauWfWJOmrT+yJpvdnPuD\nObiNomyk20P/hIxLjDzoVSfgPUCal+qaZRQ9UMpa/HnhsheKroO1522ZbJH9ojP5\nFz2+MBWSlFjL5a8cvGcztFwDzmP0xD02wIoN6m6J4SrqReoRtMj21QIDAQABozUw\nMzAOBgNVHQ8BAf8EBAMCBaAwEwYDVR0lBAwwCgYIKwYBBQUHAwIwDAYDVR0TAQH/\nBAIwADANBgkqhkiG9w0BAQsFAAOCAQEAsdpQqpBh8n5mTilaP+BPiaDQoPq2Xvfh\nH+22hMzd7ddMmEkYkctD2THXh9DrqW/b9R2KMcx2k+E04Dy0etcvIr2utxpVTSaw\nWn62t0WhrsVC+qKRwlO34Idr9oeO2mKFRYMvfkBFXeV/eesDi1d4+rwiqhzlbbwh\nzxU5No8aKe5MqU46Dk9usyFjhO3gUonQ4yXb0aEl5juNU1OhYgJf2X4VzIy68mCM\nVDuHs02AJx2Xl9LuEfHaKK01N9I7ED9vlwcfzfpuzioEJnQigTQndr3RhnonBNiZ\nS0Q0HzZmYE7c9iK80d9n8+N5NpjJnd7IUKiQfplAqV0BhRcGV7TPtw==\n-----END CERTIFICATE-----",
        "key": "-----BEGIN RSA PRIVATE KEY-----\nMIIEpAIBAAKCAQEA7ONNZ1LqU0nYcnz+0T7DtISfMlmJBiB9+gB9g6y0Ta0p7/3J\nUl9XRR50QsdRG/HyY2D8Mw9lyvY0QL/jF14IdwjMhLc88liLx2Lj1fzUGtTujo2N\nTUj0QgXH0hDe5Y3BNbMXYsAEn2SHbK0cqoz1Xa1PN93qtOflVlEofN5g15fV0akS\ndRoBs+YRpAq/C1C6SGjpA3CDb2dHbOuds04Tkxta0Y3wSnCbbkUwSxlaJ2YPrseC\nV8/f/ZtgOAxI5PkSrL5t6xNERyxEegk3OOwtdqkrHfVJzIidYv2VxOh27pTbuaTc\nMntTd4d7iyHiUpf7Lh/QYxly4KRvkO8zuzPs/wIDAQABAoIBAQDBEYk2myeMdnVI\n6oMsu4D6EdVGTh8VtUm3hsrFlO7nGClEBo4c6sPP+8A55QVAR3OMd7FVVElcoEl6\nNGq70tIHYk5+JHIx+uUBvqF++K9Opxk/ajMuODWL0fZISaPBaEV+pNZ8j4+08VRG\nyNjCUFxy2xvjpqTyiMYsEb1z+DLpwq8BOVlKAiKqnUzge0qJnZMKiC3t+zendjzT\nOMjVbQvvowRmBarYYTgSE3PbD4xhBGQVRwyCFtjU9gKDXcmJ6/JWHvT6XgguWeD3\nRGWCjJBFqD4zPQPeVtZ43gBBqf5r+xbmSmSU19d8aBEp0WHGwH4K+5zqB/Iuca8U\nZwH4ueMxAoGBAParC9mKn3x/7qZhxi/j2189cokHOcU2lUv0LbCK6Kn6gADgoi6S\njR5DkOxccReJbvguLv75zccobWVonQ6PZKxxV/2uQ7nLgWlAPXNIe8ObUVbSum8U\nONrcOpk7u2FNjiRZ1q9vhspXi31WlDZF9xJ+HdxLcNG7/WwPoJNNOHeTAoGBAPXZ\niRI2dzNUYoK5p4jBmjL8CCxKMiwqFn0PhhJIDEyT2Q2T+FGzsJtfxZO9GmNNB/dz\nMwrEPZvlbr3tGaToROLpC37ZXamdbWDSsBGzrR/HbARbqcg2pSTFBt2Yb14poa5P\nMAAT3UoKo7tT6iBNOVVyeGlfa9ZSWN2JWVT22EBlAoGBAIa6nlpQERSbe+p89Rcr\nSaMXmPunarKBQcKeuGX5OWO+YzQXff4aJeIl3X58b3j/pBECiMDUCMWDG12MalyN\njdtyfTEegfd1ZFXstKPHL3KD30WhDBun2/YO3CLeMVbQWZcZSt2/+MuETbhEFegf\nBvie0giFXDXyjFT9pCNQxkrpAoGAa0mbf881Bbl1pAbRDeUvHcDO8EUQ0CUt2zXD\nFIdMlr6OolZIr57zG1sYJfkIYv1i1KH5TGGHYVkmksu9vTF/nmaFFb5WGYcSDIbl\nP1Rh4NRuVxiJZ04xk3JIJLw5HaadYCw+jkoc8ArkmQNt0Nrigl+KHLH/pXAaR9OA\noIt4o+ECgYBmsDBLLGJlNd2FqnvkOxjh09HDJ+hlqYYk4zDcLwHSV1zibn5vUQ5U\nRWnhUZi+nNG0W/qGjSiYH21Ot+oX1uu4QeZsUlPv5oN2G3ZwQV8HtYDfewrkP53K\n+SoI0LAD6tHLDEfGL11VhexZnue8jeDB7qhV6I6Tm15pz87s8+aCIw==\n-----END RSA PRIVATE KEY-----"
        }
    resp = rc.post('/orchestrator/' + scope, json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 201:
        orch_id = str(parsed_resp["id"])
        print("Orchestrator: "+CYELLOW+orch_name+ " with ID " +CYELLOW+orch_id +" has been created" + CEND)
    else:
        print("Error occured during orchestrator creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)

    return orch_name, orch_id

def CreateVCOrchestrator(rc, scope):
    """Create vCenter Orchestrator under a Scope
    Returns:
        Orchestrator ID and Name 
    """
    orch_name = input("\nName of the vCenter orchestrator under " + scope + " scope you want to create: "+ CEND)
    address = input("\nThe IP address for vCenter: ")
    username = input("\nThe username for vCenter: ")
    password = input("\nThe password for vCenter: ")
    port = int(input("\nThe port for vCenter: "))
    print("Building orchestrator: "+CYELLOW+orch_name+ " under Scope " +CYELLOW+scope + CEND)
    
    # Now build the orchestrator, username and password is mandatory, although they can be black
    req_payload = {
    "name": orch_name,
    "type": "vcenter",
    "hosts_list": [ 
        { "host_name": address, 
        "port_number": port}],
        "username": username,
        "password": password
        }
    resp = rc.post('/orchestrator/' + scope, json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 201:
        orch_id = str(parsed_resp["id"])
        print("Orchestrator: "+CYELLOW+orch_name+ " with ID " +CYELLOW+orch_id +" has been created" + CEND)
    else:
        print("Error occured during orchestrator creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)

    return orch_name, orch_id



# =================================================================================
# Inventory
# =================================================================================
def GetInventories(rc):
    resp = rc.get('/filters/inventories')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve show inv"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowInventories(inv):
    """
        List all the inventories filter in Tetration Appliance
        Inventory ID | Name | Scope ID | Query
        """
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['UUID', 'Name', 'Scope ID', 'Query']
        data_list = [[x['id'],
                    x['name'],
                    x['app_scope_id'],
                    x['short_query']] for x in inv ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)


def GetInventoriesId(inventories, name):
    try:
        for inv in inventories:
            if name == inv["name"]:
                print ("Here is your Inventory ID: " + inv["id"])
                return inv["id"]
            else: continue
    except:
        print(URED + "Inventory {name} not found".format(name=name))

def GetInventoriesName(inventories):
    inventoriesNames = []
    try:
        for inv in inventories: 
            inventoriesNames.append(inv["name"])
        return inventoriesNames
    except:
        print(URED + "Failed to retrieve show inv"+ CEND)  

def GetInventoriesNamewithID(inventories):
    inventoriesList = []
    headers = ['Inventory Name', 'ID', 'Scope ID']
    for inv in inventories: 
        inventoriesList.append([inv["name"] , inv["id"], inv["app_scope_id"]])
    table = columnar(inventoriesList, headers, no_borders=False)
    print(table)

def GetInventoriesNamewithIDinScope(inventories,scope_id):
    inventoriesList = []
    headers = ['Inventory Name', 'ID']
    for inv in inventories: 
        if inv["app_scope_id"] ==scope_id: inventoriesList.append([inv["name"] , inv["id"]])
    table = columnar(inventoriesList, headers, no_borders=False)
    print(table)
    
def getInventoryDetail(rc, id):
    resp = rc.get('/filters/inventories/'+ id)

    if resp.status_code != 200:
        print(URED + "Failed to retrieve inventory detail"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()  

def showInvDetail(invDetail):
    UUID = invDetail['id']
    NAME = invDetail['name']
    SCOPE_ID = invDetail['app_scope_id']
    headers = ['UUID', 'Name', 'Scope ID', 'Query']
    data_list = [[UUID,
                        NAME,
                        SCOPE_ID,
                        x] for x in invDetail['query']['filters']]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def CreateInventory(rc):
    name = input (CGREEN + "\nKey in the name of the inventory filter you want to create: "+ CEND)
    inventories = GetInventories (rc)
    for inv in inventories:
        if name == inv["name"]:
            print(URED + "\nInventory {name} is conflict with existing inventory filter. Please choose different name".format(name=name)+ CEND)
        else:
            scopes = GetApplicationScopes(rc)
            scope = input (CGREEN + "\nWhat is the scope (Root:Sub) you want your filter belong to: "+ CEND)
            scope_id = GetAppScopeId(scopes,scope)
            subnet = input (CGREEN + "\nWhat is the subnet (x.x.x.x/y) you want query your inventory: "+ CEND)
            print(CGREEN + "\nBuilding inventory: "+CYELLOW+name+ CGREEN + " under Scope " +CYELLOW+scope+ CEND)
            req_payload = {
            "app_scope_id": scope_id,
            "name": name,
            "query": {
            "type": "subnet",
            "field": "ip",
            "value": subnet
            }
            }
            resp = rc.post("/filters/inventories", json_body=json.dumps(req_payload))
            parsed_resp = json.loads(resp.content)
            if resp.status_code == 200:
                sub_scope_id = str(parsed_resp["id"])
                print(Cyan + "\nInventory: "+CYELLOW+name+ Cyan + " has been created"+ CEND)
                inventory_id = GetInventoriesId (inventories,name)
            else:
                print("Error occured during sub scope creation")
                print("Error code: "+str(resp.status_code))
                print("Content: ")
                print(resp.content)
                sys.exit(3)
            return name, inventory_id

# =================================================================================
# Annotations
# =================================================================================
def downloadAnnotation(rc):
    #Download Annotation Tags to Tetration root Scope. Sample csv: need to have IP as anchor point, can add upto 32 annotations.
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: " + CEND)
    GetRootScope(vrfs)
    root_scope_name = input(CGREEN +"\nWhat is the root scope you want to upload annotations: "+ CEND)
    file_path = "AnnotationDownload.csv"
    resp= rc.download(file_path, '/assets/cmdb/download/%s' % root_scope_name)
    if resp.status_code == 200:
        print("\nDownloaded sucessful!" + CEND)
    else:
        print("Error occured during upload annotation file")
        print("Error code: "+str(resp.status_code))
        sys.exit(3)


def uploadAnnotation(rc):
    #Upload Annotation Tags to Tetration root Scope. Sample csv: need to have IP as anchor point, can add upto 32 annotations.
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
    GetRootScope(vrfs)
    root_scope_name = input(CGREEN +"\nWhat is the root scope you want to upload annotations: "+ CEND)
    file_path = "sampleAnnotationUpload.csv"
    req_payload = [tetpyclient.MultiPartOption(key='X-Tetration-Oper', val='add')]
    resp= rc.upload(file_path, '/assets/cmdb/upload/%s' % root_scope_name, req_payload)
    if resp.status_code == 200:
        print("\nUploaded sucessful!" + CEND)
    else:
        print("Error occured during upload annotation file")
        print("Error code: "+str(resp.status_code))
        sys.exit(3)


# =================================================================================
# Scopes
# =================================================================================

def GetApplicationScopes(rc):
    resp = rc.get('/app_scopes')

    if resp.status_code != 200:
        print("Failed to retrieve app scopes")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetAppScopeId(scopes,name):
    try:
        return [scope["id"] for scope in scopes if scope["name"] == name][0]
    except:
        print("App Scope {name} not found".format(name=name))        

def ShowScopes(scopes):
    """
        List all the Scopes in Tetration Appliance
        Scope ID | Scope Name | Parent Scope | VRF | Policy Priority
        """
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Number', 'Scope ID', 'Name', 'Parent Scope', 'VRF', 'Policy Priority']
        data_list = [[i+1, x['id'],
                    x['name'],
                    x['parent_app_scope_id'],
                    x['vrf_id'], x['policy_priority']] for i,x in enumerate(scopes) ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)


def GetAppScopeName(scopes,id):
    try:
        return [scope["name"] for scope in scopes if scope["id"] == id][0]
    except:
        print("App Scope {id} not found".format(name=name)) 

def ShowApplicationScopes(scopes):
    """
        List all the Scopes in Tetration Appliance
        Scope ID | Name | Policy Priority | Query | VRF ID | Parent Scope ID | Root Scope ID | Created At | Updated At
        """
    headers = ['Scope ID', 'Name', 'Policy Priority', 'Query', 'VRF ID', 'Parent Scope ID', 'Root Scope ID', 'Created At', 'Updated At']
    data_list = []
    for x in scopes: data_list. append([x['id'],
                    x['name'],
                    x['policy_priority'],
                    x['short_query'],
                    x['vrf_id'],
                    x['parent_app_scope_id'],
                    x['root_app_scope_id'],
                    time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(x['created_at'])),
                    time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(x['updated_at']))])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetVRFs(rc):
    # Get all VRFs in the cluster
    resp = rc.get('/vrfs')

    if resp.status_code != 200:
        print("Failed to retrieve app scopes")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowVRFs(vrfs):
    """
        List all the Apps in Tetration Appliance
        VRF ID | Created At | Updated At | Name | Tenant name | Root Scope ID
        """
    data_list = []
    headers = ['VRF ID', 'Created At', 'Updated At', 'Name', 'Tenant Name', 'Root Scope ID']
    for x in vrfs: 
        data_list.append([x['id'], time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(x['created_at'])), time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(x['updated_at'])), x['name'], x['tenant_name'], x['root_app_scope_id']]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetRootScope(vrfs):
    #return list of Root Scopes and its' names
    rootScopes = []
    headers = ['Root Scope Name', 'VRF ID']
    for vrf in vrfs:
        rootScopes.append([vrf["name"] , vrf["vrf_id"]])
    table = columnar(rootScopes, headers, no_borders=False)
    print(table)

def GetAllSubScopeNames(scopes, name):
    subScopeNames = []
    try:
        for scope in scopes: 
            if name in scope["name"]:
                subScopeNames.append(scope["name"])
            else: continue
        return subScopeNames
    except:
        print(URED + "App Scope {name} not found".format(name=name))

def build_root(rc):
    """Build new root scope if required. ie if not existing in validate_current function.
    Returns:
        root_app_scope_id: App scope id as created for the root scope
    """
    root_scopes = defaultdict(str)
    root_ids = defaultdict(str)
    root_app_scope_id = ""

    resp = rc.get("/vrfs")
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        for vrf in parsed_resp:
            root_scopes[vrf["name"]] = vrf["id"]
            root_ids[vrf["name"]] = vrf["root_app_scope_id"]
    else:
        print("No root scopes have been defined.")

    root_scope = input("Enter the name of the new root scope: ")
    print("Building root scope: "+CYELLOW+root_scope+CEND)
    root_scope_id = input("Enter the root scope id: ")
    for scope_id in root_scopes.items():
        if scope_id == root_scope_id:
            print("This id is already in use.  Please try again with a unique root scope id.")
            sys.exit(0)

# Now build the root scope
    req_payload = {
        "id": root_scope_id,
        "name": root_scope
    }

    resp = rc.post("/vrfs", json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)

    if resp.status_code == 200:
        print("Root scope "+root_scope+" created with id "+str(parsed_resp["id"]))
        root_app_scope_id = parsed_resp["root_app_scope_id"]
    else:
        print("Error occured during root scope creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)

    return root_scope, root_app_scope_id  

def defineRoot(rc):
    """Validate existing root scopes and identify target root if existing.
    If target root does not exist, build it via build_root function

    Returns:
        root_scope: Root scope name for scope build
        root_app_scope_id: App scope id for root scope
    """

    root_scopes = defaultdict(str)
    root_ids = defaultdict(str)
    root_app_scope_id = ""

    resp = rc.get("/vrfs")
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        for vrf in parsed_resp:
            root_scopes[vrf["name"]] = vrf["id"]
            root_ids[vrf["name"]] = vrf["root_app_scope_id"]
    if root_scopes:
        print("\nYou have the following "+str(len(root_scopes))+" root scopes configured.")
        print(*root_scopes, sep="\n")
        print(CGREEN)
        root_scope = input("Enter the name of the root scope to use for scope definition: "+CEND)

        if root_scope in root_scopes:
            print("Using root scope "+root_scope+" with id "+str(root_scopes[root_scope])+
                  " for scope definition")
            root_app_scope_id = str(root_ids[root_scope])
        else:
            print("That root scope does not exist.")
            while True:
                response = input("Would you like to create a new root scope? [y/n]: ").lower()
                if response == "n":
                    print("Okay.  Please validate the inputs and try again.")
                    sys.exit(0)
                elif response == "y":
                    root_scope, root_app_scope_id = build_root(rc)
                    break
                else:
                    print("Invalid entry, please try again...")
    else:
        print("No root scopes have been defined.")

        while True:
            response = input("Would you like to define a new root scope? [y/n]: ").lower()
            if response == "n":
                print("Okay.  Please validate the inputs and try again.")
                sys.exit(0)
            elif response == "y":
                root_scope, root_app_scope_id = build_root(rc)
                break
            else:
                print("Invalid entry, please try again...")

    return root_scope, root_app_scope_id  


def build_subscope(rc):
    """Build sub scope under root scope

    Returns:
        sub_scope: Sub scope name for scope build
        sub_cope_id: App scope id for sub scope
    """
    root_scope, root_scope_vrf_id = defineRoot(rc)
    scopes = GetApplicationScopes(rc)
    root_scope_id = GetAppScopeId(scopes,root_scope)
    sub_scope = input("Name of the sub scope under Root Scope " + root_scope + " you want to create: ")
    subnet = input("Which subnet or IP you want your query is (X.X.X.X/Y): ")
    print("Building sub scope: "+CYELLOW+sub_scope+ " under Root Scope " +CYELLOW+root_scope)
    
    # Now build the sub scope
    req_payload = {
        "short_name": sub_scope,
            "short_query": {
                "type": "subnet",
                "field": "ip",
                "value": subnet
            },
        "parent_app_scope_id": root_scope_id
    }
    
    resp = rc.post("/app_scopes", json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 200:
        sub_scope_id = str(parsed_resp["id"])
        print("Sub scope: "+CYELLOW+sub_scope+ "with scope ID " +CYELLOW+sub_scope_id +" has been created")
    else:
        print("Error occured during sub scope creation")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)

    return sub_scope, sub_scope_id


def commit_scopes(rc):
    #Commit scope changes
    scopes = GetApplicationScopes(rc)
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: " + CEND)
    GetRootScope(vrfs)
    scope = input (CGREEN + "\nWhat is the Root scope you want your filter belong to: " + CEND)
    scope_id = GetAppScopeId(scopes,scope)
    
    # commit scope changes
    req_payload = {
        "root_app_scope_id": scope_id
    }
    
    resp = rc.post("/app_scopes/commit_dirty", json_body=json.dumps(req_payload))
    parsed_resp = json.loads(resp.content)
    if resp.status_code == 202:
        print("Scope commits are undergoing ")
    else:
        print("Error occured during scope changes commit")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)
  


# =================================================================================
# Agents
# =================================================================================

def GetSensors(rc):
    resp = rc.get('/sensors')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve sensors list")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowAgents(sensors):
    """
        List all the agents registered in Tetration Appliance
        Hostname | Agent Type | Last checkin | Install Date | Version | Scopes
        """
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Number', 'Host Name', 'UUID', 'Agent Type', 'Last Check-in', 'Install Date', 'Version', 'Scopes']
        data_list = [[i+1, x['host_name'], x['uuid'], x['agent_type'], time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(x['last_config_fetch_at'])), time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(x['created_at'])), x['current_sw_version'], ','.join(set([y['vrf'] for y in x['interfaces']])) ]for i,x in enumerate(sensors['results']) ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetAgentProfiles(rc):
    resp = rc.get('/inventory_config/profiles')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve sensors list" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowAgentProfiles(profiles):
    """
        List all the Agent Profiles in Tetration Appliance
        Agent Profile ID | Name | Auto Upgrade | PID Lookup | Enforcement Disabled | Forensics | Meltdown | SideChannel
        """
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Agent Profile ID', 'Name', 'Auto Upgrade Disabled', 'PID Lookup', 'Enforcement Disabled', 'Forensics', 'Meltdown', 'SideChannel' ]
        data_list = [[x['id'],
                    x['name'], x['auto_upgrade_opt_out'], x['enable_pid_lookup'], x['enforcement_disabled'], x['enable_forensics'], x['enable_meltdown'], x['enable_cache_sidechannel'] ] for x in profiles ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)


def GetProfilesNamewithID(Profiles):
    ProfilesList = []
    try:
        for prof in Profiles: 
            ProfilesList.append([prof["name"] , prof["id"]])
        return ProfilesList
    except:
        print(URED + "Failed to retrieve Profiles list"+ CEND) 

def GetProfilesId(Profiles, name):
    try:
        for prof in Profiles:
            if name == prof["name"]:
                print (Cyan +"Here is your Profile ID: " + prof["id"])
                return prof["id"]
            else: continue
    except:
        print(URED + "Profile {name} not found".format(name=name)+ CEND)

def getProfileDetail(rc, id):
    resp = rc.get('/inventory_config/profiles/'+ id)

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Profile detail"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowProfileDetail(details):
    """
        List all the Agent Profile detail in Tetration Appliance
        Agent Profile ID | Name | Auto Upgrade | PID Lookup | Enforcement Disabled | Forensics | Meltdown | SideChannel
        """
    data_list = []
    headers = ['Agent Profile ID', 'Name', 'Auto Upgrade Disabled', 'PID Lookup', 'Enforcement Disabled', 'Forensics', 'Meltdown', 'SideChannel' ]
    data_list.append([details['id'],
                    details['name'], details['auto_upgrade_opt_out'], details['enable_pid_lookup'], details['enforcement_disabled'], details['enable_forensics'], details['enable_meltdown'], details['enable_cache_sidechannel']])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetConfigIntents(rc):
    resp = rc.get('/inventory_config/interface_intents')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve sensors list" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowConfigIntents(intents):
    """
        List all the Config Intents in Tetration Appliance
        VRF ID | VRF Name | Filter ID | Filter Name
        """
    data_list = []
    headers = ['VRF ID', 'VRF Name', 'Filter ID', 'Filter Name']
    for intent in intents.values():
        data_list.append([intent[0]['vrf_id'] , intent[0]['vrf_name'], intent[0]['inventory_filter']['id'], intent[0]['inventory_filter']['name']])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def getOSPlatform(rc):
    resp = rc.get('/sw_assets/platforms')

    if resp.status_code != 200:
        print(CRED + "Failed to retrieve OS Platforms supported list")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowOSPlatform(OSs):
    """
        List all the OS Platforms supported in Tetration Appliance
        Agent Type | Platform | Architecture
        """
    headers = ['Platform', 'Agent Type', 'Architecture']
    ImportList = []
    OSList = []
    for key,value in OSs.items() :
        ImportList.append(value)
    for list in ImportList:
        for ele in list: OSList.append([ele["platform"], ele["agent_type"],ele["arch"]])
    table = columnar(OSList, headers, no_borders=False)
    print(table)

def getOSPlatform(rc):
    resp = rc.get('/sw_assets/platforms')

    if resp.status_code != 200:
        print(CRED + "Failed to retrieve OS Platforms supported list")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetOSVersion(OSes, name):
    ImportList = []
    VersionList = []
    for key,value in OSes.items() :
        ImportList.append(value)
    try:
        for list in ImportList:
            for ele in list:
                if name in ele["platform"]:
                    if ele["platform"] in VersionList: continue
                    else: VersionList.append(ele["platform"])
                else: continue
        return VersionList
    except:
        print(CRED + "Platform {name} not found".format(name=name))

def GetConfigIntents(rc):
    resp = rc.get('/inventory_config/interface_intents')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve sensors list" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowConfigIntents(intents):
    """
        List all the Config Intents in Tetration Appliance
        VRF ID | VRF Name | Filter ID | Filter Name
        """
    data_list = []
    headers = ['VRF ID', 'VRF Name', 'Filter ID', 'Filter Name']
    for intent in intents.values():
        data_list.append([intent[0]['vrf_id'] , intent[0]['vrf_name'], intent[0]['inventory_filter']['id'], intent[0]['inventory_filter']['name']])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def DownloadAgent(rc):
    OSes = getOSPlatform(rc)
    print (CGREEN+ "\nHere is the OS Supported in this release: SUSE, CentOS, RedHat, OracleServer, Ubuntu, MS, AIX")
    OS = input (CYELLOW + "\nWhich OS you want to download with detail version supported: ")
    if ((OS == "windows") or (OS == "Windows")): 
        VersionList = GetOSVersion (OSes, "MS")
    else: VersionList = GetOSVersion (OSes, OS)
    print (CGREEN + "\nHere are the versions that are supported in this release: ")
    print(*VersionList, sep = "\n")
    platform = input (CYELLOW + "\nWhich platform and version (copy and paste here) you want to download from above: ")
    types = input (CYELLOW + "\nWhich type of sensors (sensor or enforcer) you want to download: " )
    arch = input (CYELLOW + "\nWhich arch of sensors (x86_64 or s390x) you want to download: " )
    print("tet-sensor-3.3.2.12-1."+platform+ "-tet." + types + "." + arch)
    file_name = str("tet-sensor-3.3.2.12-1."+platform+ "-tet." + types + "." + arch)
    file_name = file_name + ".rpm"
    resp = rc.download(file_name, "/sw_assets/download?platform=" + platform + "&agent_type=" + types + "&arch=" + arch)

    if resp.status_code == 200:
        print("\nDownloaded sucessful!" + CEND)
    else:
        print("Error occured during download hash")
        print("Error code: "+str(resp.status_code))
        sys.exit(3)


def Createprofile(rc):
    name = input (CYELLOW + "\nKey in the name of the Agent Config Profile you want to create: "+ CEND)
    profiles = GetAgentProfiles (rc)
    for prof in profiles:
        if name == prof["name"]:
            print(URED + "\nProfile {name} is conflict with existing Agent Config profile. Please choose different name".format(name=name) + CEND)
        else:
            scopes = GetApplicationScopes(rc)
            vrfs = GetVRFs(rc)
            print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
            GetRootScope(vrfs)
            root_scope = input (CYELLOW + "\nWhich Root Scope above you want to get your orchestrator: "+ CEND)
            scope_id = GetAppScopeId(scopes,root_scope)
            print (CGREEN + "Here is some config you need to define: "+ CEND)
            
            auto_upgrade = input (CYELLOW + "Auto upgrade agents (y/n?): ")
            if auto_upgrade == 'y': auto_upgrade = False
            else: auto_upgrade = True
            
            pid_lookup = input (CYELLOW + "Allow Process ID Lookup (y/n?): ")
            if pid_lookup == 'y': pid_lookup = True
            else: pid_lookup = False

            enforcement = input (CYELLOW + "Allow Agent enforcement - Host Based FW (y/n?): ")
            if enforcement == 'y': enforcement = False
            else: enforcement = True

            forensics = input (CYELLOW + "Enable Secure Forensics Security Events for servers (y/n?): ")
            if forensics == 'y': forensics = True
            else: forensics = False

            meltdown = input (CYELLOW + "Enable Meltdown detection for servers (y/n?): ")
            if meltdown == 'y': meltdown = True
            else: meltdown = False

            sidechannel = input (CYELLOW + "Enable SideChannel Attack detection for servers (y/n?): ")
            if sidechannel == 'y': sidechannel = True
            else: sidechannel = False

            print(CGREEN + "\nBuilding profile: "+CYELLOW+name+ CGREEN + " under Scope " +CYELLOW+root_scope + CEND)
            req_payload = {
            "name": name,
            "root_app_scope_id": scope_id,
            "data_plane_disabled": False,
            "auto_upgrade_opt_out": auto_upgrade,
            "enable_pid_lookup": pid_lookup,
            "enforcement_disabled": enforcement,
            "enable_forensics": forensics,
            "enable_meltdown": meltdown,
            "enable_cache_sidechannel": sidechannel,
            "allow_broadcast": True,
            "allow_multicast": True,
            "allow_link_local": True
            }
            resp = rc.post("/inventory_config/profiles", json_body=json.dumps(req_payload))
            parsed_resp = json.loads(resp.content)
            if resp.status_code == 200:
                sub_scope_id = str(parsed_resp["id"])
                print(Cyan + "\nProfile: "+CYELLOW+name+ Cyan + " with ID" + parsed_resp["id"]+" has been created"+ CEND)
            else:
                print("Error occured during Agent Config Profile creation")
                print("Error code: "+str(resp.status_code))
                print("Content: ")
                print(resp.content)
                sys.exit(3)
            return parsed_resp["name"], parsed_resp["id"]

def ApplyProfile2Filter(rc, profile_id):
    #Apply Agent profile into filter
    filter_id =""
    choice = input (CYELLOW + "\nDo you want to apply your profile to Scope (S) or Filter (F)? ")
    if choice == ("Scope" or "scope" or "s" or "S"):
        scope_choice = input (CYELLOW + "\nDo you want to apply your profile to Root Scope (R) or sub scope (S)? ")
        if scope_choice == ("R" or "r"):
            vrfs = GetVRFs(rc)
            scopes = GetApplicationScopes(rc)
            print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: ")
            GetRootScope(vrfs)
            root_scope_name = input(CGREEN +"\nWhat is the root scope you want to apply your Agent Profile to? ")
            filter_id = GetAppScopeId(scopes,root_scope_name)
        if scope_choice == ("S" or "s"):
            scope_name = input(CGREEN +"\nWhat is the sub scope (Root:Subscope) you want to apply your Agent Profile to? ")
            scopes = GetApplicationScopes(rc)
            filter_id = GetAppScopeId(scopes,scope_name)
        print(CGREEN + "\nApplying profile with ID : "+CYELLOW+profile_id+ CGREEN + " into " +CYELLOW+filter_id + CEND)
        req_payload = {
            "inventory_config_profile_id": profile_id,
            "inventory_filter_id": filter_id
            }
        resp = rc.post("/inventory_config/intents", json_body=json.dumps(req_payload))
        parsed_resp = json.loads(resp.content)
        if resp.status_code == 200:
            Agent_Config_intent_id = str(parsed_resp["id"])
            print(Cyan + "\nAgent Config Intent with ID : "+CYELLOW+Agent_Config_intent_id+ Cyan +" has just been created"+ CEND)
        else:
            print("Error occured during apply Agent Config Profile to Filter")
            print("Error code: "+str(resp.status_code))
            print("Content: ")
            print(resp.content)
            sys.exit(3)
        return Agent_Config_intent_id
    if choice == ("Filter" or "filter" or "f" or "F"):
        inventories = GetInventories(rc)
        print (CGREEN + "\nHere are the names and ID of all inventories in your cluster: ")
        GetInventoriesNamewithID(inventories)
        inv_name = input (CYELLOW + "\nWhich inventory name you want to apply your agent profile: ")
        filter_id = GetInventoriesId(inventories, inv_name)
        print(CGREEN + "\nApplying profile with ID : "+CYELLOW+profile_id+ CGREEN + " into " +CYELLOW+filter_id + CEND)
        req_payload = {
            "inventory_config_profile_id": profile_id,
            "inventory_filter_id": filter_id
            }
        resp = rc.post("/inventory_config/intents", json_body=json.dumps(req_payload))
        parsed_resp = json.loads(resp.content)
        if resp.status_code == 200:
            Agent_Config_intent_id = str(parsed_resp["id"])
            print(Cyan + "\nAgent Config Intent with ID : "+CYELLOW+Agent_Config_intent_id+ Cyan +" has just been created"+ CEND)
        else:
            print("Error occured during apply Agent Config Profile to Filter")
            print("Error code: "+str(resp.status_code))
            print("Content: ")
            print(resp.content)
            sys.exit(3)
        return Agent_Config_intent_id

def remoteVRF(rc):
    #This endpoint is used to specify criteria for VRF tagging for hosts based on their source IP and source port as seen by Tetration appliance.
    vrfs = GetVRFs(rc)
    print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
    GetRootScope(vrfs)
    vrf_id = input (CGREEN + "\nWhich VRF_ID above you want to organize your telemetry: ")
    src_subnet = input (CGREEN + "\nWhat is the source subnet (Ex. 192.168.1.0/24): ")
    src_port_range_start = input (CGREEN + "\nWhat is the source port range start(Ex. 0): ")
    src_port_range_end = input (CGREEN + "\nWhat is the source port range end (Ex. 65535): ")
    print(CGREEN + "\nMoving your telemetry to VRF " + vrf_id + CEND)
    req_payload = {
    "src_subnet": src_subnet,
    "src_port_range_start": int(src_port_range_start),
    "src_port_range_end": int(src_port_range_end),
    "vrf_id": int(vrf_id)}


    resp = rc.post("/agentnatconfig", json_body=json.dumps(req_payload))
    if resp.status_code == (201 or 200):
        print(Cyan + "Your telemetry has been moved to VRF " + vrf_id + CEND)
    else:
        print("Error occured during moving telemetry to VRF")
        print("Error code: "+str(resp.status_code))
        print("Content: ")
        print(resp.content)
        sys.exit(3)



def main():
    print (BLINK + BOLD+ CGREEN + "Welcome to Tetration CLI !!!" + CEND)
    print (BOLD+ Cyan+ UNDERLINE + ITALIC +"Object support:"+ CEND)
    print (BOLD+ CYELLOW +"- agents \n- inventories \n- vrfs \n- applications \n- users \n- roles \n- scopes \n- annotations \n- flow \n- orchestrators \n- policies \n- report" + CEND)
    print (BOLD+ Cyan+ UNDERLINE + ITALIC +"Operator support:"+ CEND)
    print (BOLD+ CYELLOW +"- show: show all  \n- show item: show detail of object item \n- create item: create an object item \n- setup: onboard a new cluster \n- clear: delete all objects in one root scope" + CEND)
    rc = CreateRestClient()
    command = input ("tetcli #  "+ CEND)

    # Overall
    if command == "show" or command == "show -h" or command =="show h" or command =="show help" or command =="show ?":
        print (Cyan+ "Sub commands support: inventories, vrfs, applications, users, roles, scopes, flows, agents, orchestrators, policies "+ CEND)

    # Agents
    if command == "agents" or command =="agents -h" or command =="agents h" or command =="agents help" or command =="agents ?":
        print (Cyan+ "Sub commands support: show, profiles, intents, download "+ CEND)
    if command == "show agents h" or command =="show agents help" or command =="show agents ?": 
        print (Cyan+ "Items support: all, os , osversion, profiles, intents "+ CEND)
    if command == "show agents all": 
        sensors = GetSensors(rc)
        print (CGREEN + "Here is the sensors detail: " + CEND)
        ShowAgents(sensors)
    if command == "show agents os":
        OS = getOSPlatform(rc)
        ShowOSPlatform(OS)
    if command == "show agents osversion":
        OS = getOSPlatform(rc)
        print (CGREEN+ "\nHere is the OS Supported in this release: SUSE, CentOS, RedHat, OracleServer, Ubuntu, MS, AIX"+ CEND)
        platform = input (CYELLOW + "\nWhich OS you want to get more detail version supported: "+ CEND)
        if ((platform == "windows") or (platform == "Windows")): 
            VersionList = GetOSVersion (OS, "MS")
        else: VersionList = GetOSVersion (OS, platform)
        print (CGREEN + "\nHere are the versions that are supported in this release: " + CEND)
        print(*VersionList, sep = "\n")

    # Agent Profiles
    if command == "show agents profiles h" or command =="show agents profiles help" or command =="show agents profiles ?" or command =="show agents profiles": 
        print (Cyan+ "Items support: all, detail "+ CEND)
    if command == "show agents profiles all": 
        profiles = GetAgentProfiles(rc)
        print (CGREEN + "Here is the all Agent Config Profiles configured in your cluster: " + CEND)
        ShowAgentProfiles(profiles)
    if command == "show agents profiles detail":
        Profiles = GetAgentProfiles(rc)
        ProfilesList = GetProfilesNamewithID(Profiles)
        print (CGREEN + "\nHere are the names and ID of all Profiles in your cluster: " + CEND)
        print(*ProfilesList, sep = "\n")
        prof_name = input (CYELLOW + "\nWhich Profile name you want to get more detail: "+ CEND)
        prof_id = GetProfilesId(Profiles, prof_name)
        prof_detail = getProfileDetail(rc, prof_id)
        print (CGREEN + "\nHere are the detail of your profile: " + CYELLOW + prof_name + CEND)
        ShowProfileDetail(prof_detail)
    if command == "agents profiles create":
        profile_id = Createprofile (rc)[1]
        ApplyProfile2Filter (rc, profile_id)

    # Agent Config Intents
    if command == "show agents intents h" or command =="show agents intents help" or command =="show agents intents" or command =="show agents intents ?": 
        print (Cyan+ "Items support: all "+ CEND)
    if command == "show agents intents all": 
        intents_list = GetConfigIntents(rc)
        print (CGREEN + "Here is the all Agent Config Intents configured in your cluster (Apply VRF to Filter): " + CEND)
        ShowConfigIntents(intents_list)

     # Agent Download
    if command == "agents download h" or command =="agents intents help" or command =="agents download" or command =="agents download ?": 
        print (Cyan+ "Download Tetration Installation file. Items support: none "+ CEND)
    if command == "agents download":
        agent = DownloadAgent(rc)

    # Scopes show
    if command == "scopes" or command =="scopes -h" or command =="scopes h" or command =="scopes help" or command =="scopes ?":
        print (Cyan+ "Sub commands support: show, create"+ CEND)
    if command == "show scopes h" or command =="show scopes help" or command =="show scopes" or command =="show scopes ?": 
        print (Cyan+ "Items support: all, roots, subscopes  "+ CEND)
    if command == "show scopes all": 
        scopes = GetApplicationScopes(rc)
        print (CGREEN + "Here is all scopes in your cluster: " + CEND)
        ShowApplicationScopes(scopes)
    if command == "show scopes roots":
        vrfs = GetVRFs(rc)
        print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
        GetRootScope(vrfs)
    if command == "show scopes subscopes" or command =="show scopes sub":
        scopes = GetApplicationScopes(rc)
        scope_name = input(CGREEN + "Enter the name of the scope (RootScope:SubScope) you want to get subscopes: "+CEND)
        try:
            scope_id = GetAppScopeId(scopes,scope_name)
            print(CGREEN + "\nHere is your Root Scope ID for your root scope "+ Cyan + scope_id+ CEND)
            print (CGREEN + "\nIt has below Sub Scopes including itself: "+ CEND)
            subScopeNames = GetAllSubScopeNames(scopes, scope_name)
            for elem in subScopeNames:
                print (elem)
        except:
            print(URED + "Your root Scope is not exist. Try again"+ CEND)


    # Scopes create
    if command == "scopes create h" or command =="scopes create help" or command =="scopes create" or command =="scopes create ?": 
        print (Cyan+ "Items support: root, subscope, commit  "+ CEND)
    if command == "scopes create root": 
        build_root(rc)
    if command == "scopes create subscope" or command =="scopes create sub":
        build_subscope(rc)
    if command == "scopes create commit" or command =="scopes create com":
        commit_scopes(rc)

    # vrfs show
    if command == "vrfs" or command =="vrf -h" or command =="vrfs h" or command =="vrfs help" or command =="vrfs ?":
        print (Cyan+ "Sub commands support: show, create"+ CEND)
    if command == "show vrfs h" or command =="show vrfs help" or command =="show vrfs" or command =="show vrfs ?": 
        print (Cyan+ "Items support: all "+ CEND)
    if command == "show vrfs all": 
        vrfs = GetVRFs(rc)
        print (Cyan + "\nHere are list of VRFs in your cluster: "+ CEND)
        ShowVRFs(vrfs)

    # VRFs create
    if command == "vrfs create h" or command =="vrfs create help" or command =="vrfs create" or command =="vrfs create ?": 
        print (Cyan+ "Items support: remote  "+ CEND)
    if command == "vrfs create remote": 
        remoteVRF(rc)

    # Annotations
    if command == "annotations h" or command =="annotations help" or command =="annotations" or command =="annotations ?" or command == "anno h" or command =="anno help" or command =="anno" or command =="anno ?": 
        print (Cyan+ "Items support: download, upload  "+ CEND)
    if command == "annotations download" or command =="annotations down" or command == "anno download" or command =="anno down": 
        downloadAnnotation(rc)
    if command == "annotations upload" or command =="annotations up" or command == "anno upload" or command =="anno up":
        uploadAnnotation(rc)


    # inventories show
    if command == "inventories" or command =="inventories -h" or command =="inventories h" or command =="inventories help" or command =="inventories ?":
        print (Cyan+ "Sub commands support: show, create"+ CEND)
    if command == "show inv h" or command =="show inv help" or command =="show inv" or command =="show inv ?": 
        print (Cyan+ "Items support: all, detail  "+ CEND)
    if command == "show inv all": 
        inv = GetInventories(rc)
        print (CGREEN + "Here is the all inventories configured in your cluster: " + CEND)
        ShowInventories(inv)
    if command == "show inv detail" or command =="show inv de" or command =="show inv d" or command =="show inv det":
        inventories = GetInventories(rc)
        print (CGREEN + "\nHere are the names and ID of all inventories in your cluster: "+ CEND)
        GetInventoriesNamewithID(inventories)
        inv_name = input (CYELLOW + "\nWhich inventory name you want to get more detail: "+ CEND)
        inv_id = GetInventoriesId(inventories, inv_name)
        inv_detail = getInventoryDetail(rc, inv_id)
        print (Cyan + "\nHere is the detail of your inventory " + inv_name + " :" + CEND)
        showInvDetail(inv_detail)


    # Inventory create
    if command == "inventories create h" or command =="inventories create help" or command =="inventories create" or command =="inventories create ?" or command == "inv create h" or command =="inv create help" or command =="inv create" or command =="inv create ?": 
        print (Cyan+ "Items support: none  "+ CEND)
    if command == "inventories create": 
        CreateInventory (rc)

     # orchestrators show
    if command == "orchestrators" or command =="orchestrators -h" or command =="orchestrators h" or command =="orchestrators help" or command =="orchestrators ?":
        print (Cyan+ "Sub commands support: show, create"+ CEND)
    if command == "show orc h" or command =="show orc help" or command =="show orc" or command =="show orc ?": 
        print (Cyan+ "Items support: all, detail  "+ CEND)
    if command == "show orc all": 
        orchestrators = GetOrchestrators(rc)
        print(CGREEN + "Here is the list of Orchestrators in your cluster: " + CEND)
        ShowOrchestrators(orchestrators)
    if command == "show orc detail" or command =="show orc de" or command =="show orc d" or command =="show orc det":
        vrfs = GetVRFs(rc)
        print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
        GetRootScope(vrfs)
        scope_name = input (CYELLOW + "\nWhich Root Scope you want to get your orchestrator: "+ CEND)
        Orchestrators = GetOrchestratorsScope(rc,scope_name)
        GetOrchestratorsNamewithID(Orchestrators)
        orc_name = input (CYELLOW + "\nWhich Orchestrator name you want to get more detail: " + CEND)
        orc_id = GetOrchestratorsId(Orchestrators, orc_name)
        orc_detail = getOrchestratorDetail(rc, scope_name, orc_id)
        ShowOrchestrator(orc_detail)

    # Orchestrator create
    if command == "orchestrators create h" or command =="orchestrators create help" or command =="orchestrators create" or command =="orchestrators create ?" or command == "orc create h" or command =="orc create help" or command =="orc create" or command =="orc create ?": 
        print (Cyan+ "Items support: vcenter, k8s  "+ CEND)
    if command == "orchestrators create vcenter" or command =="orchestrators create vc": 
        scopes = GetApplicationScopes(rc)
        vrfs = GetVRFs(rc)
        print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
        GetRootScope(vrfs)
        scope = input ("Which root scope above you want to create your orchestrator: "+ CEND)
        CreateVCOrchestrator(rc,scope)
    if command == "orchestrators create k8s": 
        scopes = GetApplicationScopes(rc)
        vrfs = GetVRFs(rc)
        print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
        GetRootScope(vrfs)
        scope = input ("Which root scope above you want to create your orchestrator: "+ CEND)
        CreateK8sOrchestrator(rc,scope)

    # flows show
    if command == "flows" or command =="flows -h" or command =="flows h" or command =="flows help" or command =="flows ?":
        print (Cyan+ "Sub commands support: show"+ CEND)
    if command == "show flows h" or command =="show flows help" or command =="show flows ?": 
        print (Cyan+ "Items support: dimensions, metrics  "+ CEND)
    if command == "show flows dimensions" or command =="show flows dim": 
        dimensions = GetFlowDimensions(rc)
        with open('dimensions.json', 'w') as outfile:
            json.dump(dimensions, outfile, indent=4)
        print (json.dumps(dimensions, indent=4, sort_keys=True)+ CEND)
    if command == "show flows metrics" or command =="show flows met": 
        metrics = GetFlowMetrics(rc)
        with open('metrics.json', 'w') as outfile:
            json.dump(metrics, outfile, indent=4)
        print (json.dumps(metrics, indent=4, sort_keys=True)+ CEND)


    # roles show
    if command == "roles" or command =="roles -h" or command =="roles h" or command =="roles help" or command =="roles ?":
        print (Cyan+ "Sub commands support: show, create, apply"+ CEND)
    if command == "show roles h" or command =="show roles help" or command =="show roles" or command =="show roles ?": 
        print (Cyan+ "Items support: all, detail  "+ CEND)
    if command == "show roles all": 
        roles = getRoles(rc)
        print (CGREEN + "Here is the all roles configured in your cluster: " + CEND)
        ShowRoles(roles)
    if command == "show roles detail" or command =="show roles de" or command =="show roles d" or command =="show roles det":
        roles = getRoles(rc)
        scopes = GetApplicationScopes(rc)
        print (CGREEN + "\nHere are the names and ID of all roles in your cluster: "+ CEND)
        GetRolesNamewithID(roles)
        role_name = input (CYELLOW + "\nWhich role name you want to get more detail: "+ CEND)
        role_id = GetRoleId(roles, role_name)
        role_detail = getRoleDetail(rc, role_id)
        if role_detail["app_scope_id"] =="": print (Cyan + "\nYour role " +CYELLOW+role_name+ CGREEN + " is a system provider role"+ CEND)
        else: 
            scope_id = role_detail["app_scope_id"]
            scope_name = GetAppScopeName(scopes,scope_id)
            print (Cyan + "\nHere is the detail of your Role " + role_name + " :"+ CEND)
            ShowRoleDetails(role_detail)
            print (Cyan + "\nYour role " +CYELLOW+role_name+ CGREEN + " belongs to scope " +CYELLOW+scope_name+ CEND)



    # users show
    if command == "users" or command =="users -h" or command =="users h" or command =="users help" or command =="users ?":
        print (Cyan+ "Sub commands support: show, create"+ CEND)
    if command == "show users h" or command =="show users help" or command =="show users" or command =="show users ?": 
        print (Cyan+ "Items support: all, detail  "+ CEND)
    if command == "show users all": 
        Users = getUsers(rc)
        Roles = getRoles(rc)
        print (CGREEN + "Here is the all users configured in your cluster: " + CEND)
        ShowUsers(Users, Roles)
    if command == "show users detail" or command =="show users de" or command =="show users d" or command =="show users det":
        Users = getUsers(rc)
        Roles = getRoles(rc)
        scopes = GetApplicationScopes(rc)
        print (CGREEN + "\nHere are the details of all Users in your cluster: "+ CEND)
        ShowUsers(Users, Roles)
        User_email = input (CYELLOW + "\nWhich User email you want to get more detail: "+CEND)
        User_id = GetUserId(Users, User_email)
        User_detail = getUserDetail(rc, User_id)
        print (Cyan + "\nHere is the detail of your User " + User_email + " :"+CEND)
        ShowUserDetail(User_detail)
        scope_name = GetAppScopeName(scopes,User_detail['preferences']['root_app_scope_id'])
        print (Cyan + "\nYour User " + User_email + " belongs to root scope: " + scope_name+CEND)


    # role create
    if command == "roles create h" or command =="roles create help" or command =="roles create ?": 
        print (Cyan+ "Create role, sub command: none  "+ CEND)
    if command == "roles create": 
        CreateRole (rc)


     # role apply
    if command == "roles apply h" or command =="roles apply help" or command =="roles apply ?": 
        print (Cyan+ "Apply role to scope, sub command: none  "+ CEND)
    if command == "roles apply": 
        role2Scope (rc)


    # user create
    if command == "users create h" or command =="users create help" or command =="users create ?" or command == "user create h" or command =="user create help" or command =="user create ?": 
        print (Cyan+ "Create user, sub command: none  "+ CEND)
    if command == "users create": 
        CreateUser (rc)


    # filehash
    if command == "filehash h" or command =="filehash help" or command =="filehash" or command =="filehash ?": 
        print (Cyan+ "Items support: download, upload, delete  "+ CEND)
    if command == "filehash download" or command =="filehash down": 
        downloadHash(rc)
    if command == "filehash upload" or command =="filehash up":
        uploadHash(rc)
    if command == "filehash delete" or command =="filehash del":
        deleteHash(rc)

    # apps show
    if command == "apps" or command =="apps -h" or command =="apps h" or command =="apps help" or command =="apps ?":
        print (Cyan+ "Sub commands support: show, create"+ CEND)
    if command == "show apps h" or command =="show apps help" or command =="show apps" or command =="show apps ?": 
        print (Cyan+ "Items support: all, brief, detail, version, clusters, enforced  "+ CEND)
    if command == "show apps all": 
        apps = GetApps(rc)
        print (CGREEN + "Here is the all application configured in your cluster: " + CEND)
        ShowApps(apps)
    if command == "show apps detail" or command =="show apps de" or command =="show apps d" or command =="show apps det":
        Apps = GetApps(rc)
        AppsList = GetAppsNamewithID(Apps)
        print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: "+ CEND)
        print(*AppsList, sep = "\n")
        app_name = input (CYELLOW + "\nWhich App name you want to get more detail: "+ CEND)
        app_id = GetAppsId(Apps, app_name)
        app_infor = GetAppInfor(rc, app_id)
        ShowAppInfor(app_infor)
    if command == "show apps brief" or command =="show apps br" or command =="show apps b" or command =="show apps brie":
        Apps = GetApps(rc)
        AppsList = GetAppsNamewithID(Apps)
        print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: " + CEND)
        print(*AppsList, sep = "\n")
        app_name = input (CYELLOW + "\nWhich App name you want to get brief information: " + CEND)
        app_id = GetAppsId(Apps, app_name)
        app_detail = getAppDetail(rc, app_id)
        ShowAppDetail(app_detail)
    if command == "show apps versions" or command =="show apps ver" or command =="show apps v" or command =="show apps version":
        Apps = GetApps(rc)
        AppsList = GetAppsNamewithID(Apps)
        print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: "+ CEND)
        print(*AppsList, sep = "\n")
        app_name = input (CYELLOW + "\nWhich App name you want to get versions: "+ CEND)
        app_id = GetAppsId(Apps, app_name)
        app_version = GetAppVersions(rc, app_id)
        print (Cyan + "\nHere is the versions of your App " + app_name + " :" + CEND)
        ShowAppVersions(app_version)
    if command == "show apps clusters" or command =="show apps clus" or command =="show apps c" or command =="show apps cluster":
        Apps = GetApps(rc)
        AppsList = GetAppsNamewithID(Apps)
        print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: "+ CEND)
        print(*AppsList, sep = "\n")
        app_name = input (CYELLOW + "\nWhich App name you want to get more detail: "+ CEND)
        app_id = GetAppsId(Apps, app_name)
        clusters_list = GetClusters(rc, app_id)
        with open('Clusters.json', 'w') as outfile:
            json.dump(clusters_list, outfile, indent=4)
        if clusters_list ==[]: print (Cyan + "\nYour application " + app_name +" doesn't have any cluster. Please define it or start to try ADM now. " + CEND)
        else: 
            print (Cyan + "\nHere is the clusters (brief information, if you want detail, please open file: 'Clusters.json') of your App " + app_name + " :"+ CEND)
            ShowAppClusters(clusters_list)
    if command == "show apps enforced" or command =="show apps enf" or command =="show apps e" or command =="show apps enforce":
        Apps = GetApps(rc)
        enforcedApps = getEnforcedApps(Apps)
        if enforcedApps == []: print (CRED + "\nThere's no enforced apps in your cluster."+ CEND)
        else: 
            print (CGREEN + "\nHere are the enforced Apps in your cluster: "+ CEND)
            print(*enforcedApps, sep = "\n")


    # app create
    if command == "apps create h" or command =="apps create help" or command =="apps create ?": 
        print (Cyan+ "Create app workspace under a scope without policies, sub command: none  "+ CEND)
    if command == "apps create": 
        scopes = GetApplicationScopes(rc)
        scope = input (CGREEN +"Which parent Scope (RootScope:SubScope) you want to process: ")
        CreateApp(rc, scopes,scope)


    # policies show
    if command == "policies" or command =="policies -h" or command =="policies h" or command =="policies help" or command =="policies ?" or command =="pol" or command =="pol -h" or command =="pol h" or command =="pol help" or command =="pol ?":
        print (Cyan+ "Sub commands support: show, create, add, upload, download, convert"+ CEND)
    if command == "show policies h" or command =="show policies help" or command =="show policies" or command =="show policies ?" or command == "show pol h" or command =="show pol help" or command =="show pol" or command =="show pol ?": 
        print (Cyan+ "Items support: all, brief, detail  "+ CEND)
    if command == "show policies all" or command == "show pol all": 
        Apps = GetApps(rc)
        print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: "+ CEND)
        ShowApps(Apps)
        app_name = input (CYELLOW + "\nWhich App name you want to get more detail: "+ CEND)
        app_id = GetAppsId(Apps, app_name)
        app_policies = GetPolicies(rc, app_id)
        with open('AppPolicies.json', 'w') as outfile:
            json.dump(app_policies, outfile, indent=4)
        print (Cyan + "\nHere is the detail policies of your App " + app_name + " :"+ CEND)
        print (json.dumps(app_policies, indent=4, sort_keys=True)+ CEND)
    if command == "show policies detail" or command =="show policies de" or command =="show policies d" or command =="show policies det" or command == "show pol detail" or command =="show pol de" or command =="show pol d" or command =="show pol det":
        Apps = GetApps(rc)
        AppsList = GetAppsNamewithID(Apps)
        
        print (CGREEN + "\nHere are the names and ID of all Apps in your cluster: "+ CEND)
        print(*AppsList, sep = "\n")
        
        app_name = input (CYELLOW + "\nWhich App name you want to get more detail: "+ CEND)
        app_id = GetAppsId(Apps, app_name)
        
        detail = input (CYELLOW + "\nWhich type of policies (Default, Absolute, Catch_all) you want to see the detail: "+ CEND)
        try:
            if detail == "Default" or detail == "default" or detail == "def" or detail == "Def":
                default_detail = getDefaultDetail(rc, app_id)
                if default_detail == []: print (CRED +"\nThere's no Default Policy in your apps. Please define it or run ADM" +CEND)
                else: 
                    print (Cyan + "\nHere are the Default Policies of your App "+ app_name + " :"+ CEND)
                    with open('AppDefaultDetail.json', 'w') as outfile:
                        json.dump(default_detail, outfile, indent=4)
                    print (json.dumps(default_detail, indent=4, sort_keys=True))

            if detail == "Absolute" or detail == "absolute" or detail == "abs" or detail == "Abs":
                abs_detail = getAbsoluteDetail(rc, app_id)
                if abs_detail == []: print (CRED +"\nThere's no Default Policy in your apps. Please define it or run ADM" +CEND)
                else: 
                    print (Cyan + "\nHere are the Absolute Policies of your App "+ app_name + " :"+ CEND)
                    with open('AppAbsDetail.json', 'w') as outfile:
                        json.dump(abs_detail, outfile, indent=4)
                    print (json.dumps(abs_detail, indent=4, sort_keys=True)+ CEND)

            if detail == "Catch_all" or detail == "catch_all" or detail == "catch" or detail == "Catch" or detail == "Catch_All":
                catch_detail = getCatchAllDetail(rc, app_id)
                if catch_detail["action"] == "ALLOW": print (Cyan + "\nAnd your app "+ app_name + " is in Black List (Allow All by default)"+ CEND)
                else: print (Cyan + "\nAnd your app "+ app_name + " is in White List (Deny All by default)"+ CEND)

        except: print (CRED + " \n Please choose the correct option for type of policies (Default, Absolute, Catch_all)" + CEND)
    

    # Policies create
    if command == "policies create h" or command =="policies create help" or command =="policies create ?" or command == "pol create h" or command =="pol create help" or command =="pol create ?": 
        print (Cyan+ "Create Policies or Clusters under an existing application workspace, sub command: clusters, ports  "+ CEND)
    if command == "policies create ports" or command == "policies create port" or command == "pol create ports" or command == "pol create port": 
        CreatePolicy(rc)
    if command == "policies create clusters" or command == "policies create cluster" or command == "pol create clusters" or command == "pol create cluster" or command == "policies create clus" or command == "pol create clus": 
        CreateCluster(rc)


    # Policies add
    if command == "policies add h" or command =="policies add help" or command =="policies add ?" or command == "pol add h" or command =="pol add help" or command =="pol add ?": 
        print (Cyan+ "add Policies or Clusters under an existing application workspace, sub command: clusters, absolute, default, catch_all, ports  "+ CEND)
    if command == "policies add default" or command == "policies add def" or command == "pol add default" or command == "pol add def": 
        CreateDefaultPolicy(rc)
    if command == "policies add absolute" or command == "policies add abs" or command == "pol add absolute" or command == "pol add abs": 
        CreateAbsolutePolicy(rc)
    if command == "policies add catch_all" or command == "policies add catch" or command == "policies add Catch_all" or command == "policies add Catch" or command == "pol add catch_all" or command == "pol add catch" or command == "pol add Catch_all" or command == "pol add Catch": 
        CatchAll(rc)
    if command == "policies add clusters" or command == "policies add cluster" or command == "policies add clus" or command == "pol add clusters" or command == "pol add cluster" or command == "pol add clus": 
        CreateCluster(rc)
    if command == "policies add ports" or command == "policies add port" or command == "policies add p" or command == "pol add ports" or command == "pol add port" or command == "pol add p":
        AddServicePort(rc)

    # Policies convert
    if command == "policies convert h" or command =="policies convert help" or command =="policies convert ?" or command == "pol convert h" or command =="pol convert help" or command =="pol convert ?": 
        print (Cyan+ "Convert Policies into other formats, sub command: asa, csv, n9k  "+ CEND)
    if command == "policies convert csv" or command == "pol convert csv": 
        convApps2xls(rc)
    if command == "policies convert asa" or command == "pol convert asa":
        convApps2asa(rc)
    if command == "policies convert n9k" or command == "pol convert n9k":
        convApps2n9k(rc)

    # filehash
    if command == "filehash h" or command =="filehash help" or command =="filehash" or command =="filehash ?": 
        print (Cyan+ "Items support: download, upload, delete  "+ CEND)
    if command == "filehash download" or command =="filehash down": 
        downloadHash(rc)
    if command == "filehash upload" or command =="filehash up":
        uploadHash(rc)
    if command == "filehash delete" or command =="filehash del":
        deleteHash(rc)


    # Server Ports Upload and Download
    if command == "policies download" or command =="policies down" or command == "pol download" or command =="pol down": 
        serverPorts = GetServerPorts(rc)
        with open('ServerPortsDownload.json', 'w') as outfile:
            json.dump(serverPorts, outfile, indent=4)
        print (Cyan + json.dumps(serverPorts, indent=4, sort_keys=True))
    if command == "policies upload" or command =="policies up" or command == "pol upload" or command =="pol up":
        uploadServerPorts(rc)


    # onboard New Tetration Cluster
    if command == "setup ?" or command == "setup h" or command == "setup -h" or command == "setup help": 
        print(CGREEN +"Here are basic steps to fresh start a Tetration tenant"+ CEND)
    if command == "setup":
        print(CGREEN +"\nHere are basic steps to fresh start a Tetration tenant:"+ CEND)
        print(CGREEN +"\nStep 1: Creating new Tenant and Root Scope:"+ CEND)
        build_root(rc)
        print(CGREEN +"\nStep 2: Creating subscopes: "+ CEND)
        print(CGREEN +"\nStep 2a: Build Sub Scope"+ CEND)
        build_subscope(rc)
        print(CGREEN +"\nStep 2b: Commit scope changes"+ CEND)
        commit_scopes(rc)
        print(CGREEN +"\nStep 3: Upload annotation for inventories tagging: "+ CEND)
        uploadAnnotation(rc)
        print(CGREEN +"\nStep 4: Create Agent Config Profile: "+ CEND)
        print(CGREEN +"\nStep 4a: Create Agent Profile:"+ CEND)
        profile_id = Createprofile (rc)[1]
        print(CGREEN +"\nStep 4b: Create Agent Config Intent - Applying Agent Profile to Scope/Filter"+ CEND)
        ApplyProfile2Filter (rc, profile_id)
        print(CGREEN +"\nStep 4c: Move telemetry to Tenant VRF"+ CEND)
        remoteVRF(rc)

    # report 
    if command == "report h" or command =="report help" or command =="report ?": 
        print (BOLD+ CYELLOW + "Build report for Tetration, sub command: workloads or flows or apps  "+ CEND)
    if command == "report workloads" or command == "report wl" or command == "report workloads ?" or command == "report workloads h" or command == "report workloads help" or command == "report wl ?" or command == "report wl h" or command == "report wl help":
        print (BOLD+ CYELLOW + "Build report for Tetration workloads, sub command: all or detail or stats or software or vulnerabilities or processes "+ CEND)
        print (BOLD+ CYELLOW + "All - Report all installed workloads in your cluster in all scopes  "+ CEND)
        print (BOLD+ CYELLOW + "Detail - Detail Report about a specific workload  "+ CEND)
        print (BOLD+ CYELLOW + "Stats - Detail Workload communication report from time (t0) to time(t1)  "+ CEND)
        print (BOLD+ CYELLOW + "Software - Detail Installed Software Packages report for a specific workload  "+ CEND)
        print (BOLD+ CYELLOW + "Vulnerabilities - Detail Vulnerable Software Packages report for a specific workload or all workloads that match a CVE Score query. Sub: workload or all  "+ CEND)
        print (BOLD+ CYELLOW + "Processes - Detail Running processes report for a specific workload. Sub command: summary or all  "+ CEND)
    if command == "report flows" or command == "report flow" or command == "report flows ?" or command == "report flows h" or command == "report flows help" or command == "report flow ?" or command == "report flow h" or command == "report flow help":
        print (BOLD+ CYELLOW + "inv - Detail flow communication report about a subnet in a VRF from time (t0) to time(t1) "+ CEND)
        print (BOLD+ CYELLOW + "top - Top Talkers/Destination/Service report in excel for a scope from time (t0) to time(t1). Sub command: talkers, servers, sports, dports "+ CEND)
    if command == "report apps" or command == "report app" or command == "report apps ?" or command == "report apps h" or command == "report apps help" or command == "report app ?" or command == "report app h" or command == "report app help":
        print (BOLD+ CYELLOW + "Build report for Tetration Apps, sub command: policies or conversation "+ CEND)
        print (BOLD+ CYELLOW + "Policies - Report policies in xlsx format for a specific Application  "+ CEND)
        print (BOLD+ CYELLOW + "Conversation - Report conversation in xlsx format for a specific Application  "+ CEND)
    if command == "report workloads all" or command == "report wl all" or command == "report workloads a" or command == "report wl a": 
        sensors = GetSensors(rc)
        print (BOLD+ CYELLOW + "\nHere are all Software Sensors in your cluster: " + CEND)
        ShowAgents(sensors)
    if command == "report workloads detail" or command == "report wl detail" or command == "report workloads det" or command == "report wl det": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        agent = GetAgentProfile(rc,uuid)
        ShowAgentProfile(agent)
    if command == "report workloads stats" or command == "report wl stats" or command == "report workloads st" or command == "report wl st": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
        from_month = input(CYELLOW + "Month (mm)? "+CEND)
        from_day = input(CYELLOW + "Day (dd)? "+CEND)
        to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
        to_month = input(CYELLOW + "Month (mm)? "+CEND)
        to_day = input(CYELLOW + "Day (dd)? "+CEND)
        td = input(CYELLOW + "What is the granularity (day, hour or minute)? "+CEND)
        t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
        t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
        stats = GetWorkloadStats(rc,uuid,t0,t1,td)
        print ("Here is the detail communication for your agent with UUID: " + uuid + " from " + from_day + "/"+ from_month + "/"+ from_year+ " to " + to_day + "/"+ to_month + "/"+ to_year)
        ShowWorkloadStats(stats)
    if command == "report workloads software" or command == "report wl software" or command == "report workloads sw" or command == "report wl sw": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        packages = GetSwPackages(rc,uuid)
        print ("Here are all the software packages installed in your agent with UUID: " + uuid)
        ShowSwPackages(packages)
    if command == "report workloads vulnerabilities" or command == "report wl vulnerabilities" or command == "report workloads vul" or command == "report wl vul": 
        print (BOLD+ CYELLOW + "Vulnerabilities - Detail Vulnerable Software Packages report for a specific workload or all workloads that match a CVE Score query. Sub: workload or all  "+ CEND)
        print (BOLD+ CYELLOW + "Please choose workloads or all as subcommand"+ CEND)
    if command == "report workloads vulnerabilities all" or command == "report wl vulnerabilities all" or command == "report workloads vul all" or command == "report wl vul all": 
        get_inventory_cve(rc)
    if command == "report workloads vulnerabilities workloads" or command == "report wl vulnerabilities workloads" or command == "report workloads vul wl" or command == "report wl vul wl": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        vuls = GetVul(rc,uuid)
        print ("Here are all vulnerable packages installed in your agent with UUID: " + uuid)
        ShowVul(vuls)
    if command == "report workloads processes" or command == "report wl processes" or command == "report workloads proc" or command == "report wl proc": 
        print (BOLD+ CYELLOW + "Processes - Detail Running processes report for a specific workload. Sub command: summary or all  "+ CEND)
        print (BOLD+ CYELLOW + "Please choose summary or all as subcommand"+ CEND)
    if command == "report workloads processes all" or command == "report wl processes all" or command == "report workloads proc all" or command == "report wl proc all": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        proc = GetProc(rc,uuid)
        print ("Here are all long running processes in your agent with UUID: " + uuid)
        ShowProc(proc)
    if command == "report workloads processes summary" or command == "report wl processes summary" or command == "report workloads proc sum" or command == "report wl proc sum": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        proc = GetProcTree(rc,uuid)
        handle = proc['process_summary'][0]['summary'][0]['handle']
        procDetail = GetProcTreeDetail(rc,uuid, handle)
        print ("Here are  process snapshot detail in your agent with UUID: " + uuid)
        #print (json.dumps(procDetail, indent=4))
        ShowProcTreeDetail(procDetail)
    if command == "report flow inventories" or command == "report flow inv" or command == "report flows inv" or command == "report flows inventories":
        get_inventory_flow(rc)
    if command == "report flow top" or command == "report flows top" or command == "report flow top ?" or command == "report flows top ?":
        print (BOLD+ CYELLOW + "top - Top Talkers/Destination/Service report in excel for a scope from time (t0) to time(t1). Sub command: talkers, servers, cservice, dservice "+ CEND)
        print (BOLD+ CYELLOW + "Please provide Sub command: talkers, servers, sports, dports "+ CEND)
    if command == "report flow top talkers" or command == "report flows top talkers" or command == "report flow top t" or command == "report flows top t":
        get_flow_topTalkers(rc)
    if command == "report flow top servers" or command == "report flows top servers" or command == "report flow top s" or command == "report flows top s":
        get_flow_topDest(rc)
    if command == "report flow top sports" or command == "report flows top sports" or command == "report flow top sp" or command == "report flows top sp":
        get_flow_topSrcService(rc)
    if command == "report flow top dports" or command == "report flows top dports" or command == "report flow top dp" or command == "report flows top dp":
        get_flow_topDestService(rc)
    if command == "report apps policies" or command == "report app policies" or command == "report apps pol" or command == "report app pol":
        convApps2xls(rc)
    if command == "report apps conversation" or command == "report app conversation" or command == "report apps conv" or command == "report app conv":
        AllApps = GetApps(rc)
        appIDs = selectTetApps(AllApps)
        downloadConvs(rc, appIDs)
        with open('all-conversations.json') as config_file:
            ShowConversationTet(json.load(config_file))



    # clean all objects in Tetration scope
    if command == "clean ?" or command == "clean h" or command == "clean -h" or command == "clean help":
        print(BLINK + URED +"You are about to delete all objects under a scope."+ CEND)
    if command == "clean":
        print(BLINK + URED +"\nYou are about to delete all objects under a scope."+ CEND)
        choice = input (URED +"\nAre you sure (y/n)?"+ CEND)
        if choice == "y"or choice == "yes":
            print(CGREEN +"\nStarting clean up process"+ CEND)
            vrfs = GetVRFs(rc)
            print (Cyan + "\nHere are the names and VRF ID of all the root scopes in your cluster: "+ CEND)
            GetRootScope(vrfs)
            root_scope_name = input (CGREEN + "\nWhich Root Scope above you want to clean up the config: "+ CEND)
            clean(rc, root_scope_name)
        else: sys.exit(3)

if __name__ == "__main__":
    main()